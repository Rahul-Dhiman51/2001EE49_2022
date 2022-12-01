import pandas as pd
from openpyxl.styles.borders import Border, Side
from io import BytesIO
import shutil
import pathlib
import streamlit as st
import os.path

def oct_find(p,q,r):                                      
	if(p>0 and q>0 and r>0):    
		return 1
	elif(p>0 and q>0 and r<0):
		return -1
	elif(p<0 and q>0 and r>0):
		return 2
	elif(p<0 and q>0 and r<0):
		return -2
	elif(p<0 and q<0 and r>0):
		return 3
	elif(p<0 and q<0 and r<0):
		return -3
	elif(p>0 and q<0 and r>0):
		return 4
	elif(p>0 and q<0 and r<0):
		return -4


bor = Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'), bottom=Side(style='thin'))
h_c=['T','U','V','W','U Avg','V Avg','W Avg',r"U'=U-U avg",r"V'=V-V avg",r"W'=W-W avg",'Octant']

def single_file_octant(mod,file):
	import openpyxl
	from pandas import read_excel
	from openpyxl import Workbook
	from openpyxl import workbook,load_workbook
	from itertools import repeat
	from openpyxl.styles import PatternFill
	

	def wf(s,mod):							
		inputFilePath=s									
		df=read_excel(inputFilePath)									
		wb=Workbook()											
		sheet=wb.active													
		
		for i in range(11):										
			sheet.cell(row=2,column=i+1).value=h_c[i]	
		octant=[]

		u_mean=df['U'].mean()                                                 
		v_mean=df['V'].mean()
		w_mean=df['W'].mean()
		sheet.cell(row=1,column=14).value='Overall Octant Count'
		sheet.cell(row=1,column=45).value='Longest Subsequence Length'
		sheet.cell(row=1,column=49).value='Longest Subsquence Length with Range'
		sheet['E3'] = u_mean
		sheet['F3'] = v_mean
		sheet['G3'] = w_mean


		for i in df.index:
			sheet.cell(row=i+3,column=1).value=df['T'][i]
			sheet.cell(row=i+3,column=2).value=df['U'][i]
			sheet.cell(row=i+3,column=3).value=df['V'][i]
			sheet.cell(row=i+3,column=4).value=df['W'][i]
			sheet.cell(row=i+3,column=8).value=round(df['U'][i]-u_mean,3)
			sheet.cell(row=i+3,column=9).value=round(df['V'][i]-v_mean,3)
			sheet.cell(row=i+3,column=10).value=round(df['W'][i]-w_mean,3)
			sheet.cell(row=i+3,column=11).value=oct_find(df['U'][i]-u_mean,df['V'][i]-v_mean,df['W'][i]-w_mean)
			octant.append(oct_find(df['U'][i]-u_mean,df['V'][i]-v_mean,df['W'][i]-w_mean))
		
		def range_name_oct(mod=5000):
			octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}

			# dictionary creation in order to  mapping 
			dict_mapping={}                            
			# Creating opposite of dict_mapping dictionary                              
			opp_dict_mapping={}                                                      
			
			for i in range(4):                                            
				dict_mapping[i+1]=2*i+1-1                                           
				dict_mapping[-(i+1)]=2*(i+1)-1
				opp_dict_mapping[2*i+1-1]=i+1
				opp_dict_mapping[2*(i+1)-1]=-(i+1)

			# this function is used to find rank of each octant 
			def rankListFind(h_c):                                     
				temp_head_comp=h_c.copy()
				temp_head_comp.sort(reverse=True)
				res=[]

				for i in h_c:
					for j in range(8):
						if(i==temp_head_comp[j]):
							res.append(j+1)
							break
				# Return the list of ranks		
				return res                                                  
			
			# The given list is for finding the rank 1 octants
			def rank_1(h_c):                                         
				for i in range(8):
					if(h_c[i]==1):
						return opp_dict_mapping[i]

			# Finding the count of rank 1 in the rank 1 mod values of octant x
			def count_rank_1(h_c,x):                                         
				ct=0
				for i in h_c:
					if(x==i):
						ct+=1
				# Return the count		
				return ct                                                  
			
			# Storing rank list for different mod values in matrix
			mtrix_rank=[] 

			# storing the octs having rank 1 in diff. mod ranges and overall                                   
			rank1_list=[] 
			try:                                                  
				sheet=wb.active
			except:
				print("sheet activation error")	

			# Adding string 'user input' at specified place.
			sheet['M4']='Mod '+str(mod)                                           
			# For storing oct within ranges as 2-d matrix.
			oct_range_matrix=[]
			# Storing elements of 9 columns in a list                                                  
			col_list=[0]*9                                                     

			 # Storing header list in 'col_list' list
			col_list[0]='Octant ID'                                           


			for i in range(0,4):
				col_list[2*i+1]=(i+1)
				col_list[2*(i+1)]=-(i+1)


			# Appending header list in matrix
			oct_range_matrix.append(col_list)          

			# Adding header in the worksheet                                 
			for i in range(13,22):                                          
				sheet.cell(row=3,column=i+1).value=col_list[i-13]
				sheet.cell(row=3,column=i+1).border=bor
				if(i>13):
					sheet.cell(row=3,column=i+9).value='Rank Octant '+str(col_list[i-13])
					sheet.cell(row=3,column=i+9).border=bor
			sheet.cell(row=3,column=31).value='Rank1 Octant ID'
			sheet.cell(row=3,column=32).value='Rank1 Octant Name'
			sheet.cell(row=3,column=31).border=bor
			sheet.cell(row=3,column=32).border=bor

			# Resetting values in list 'col_list'
			col_list=[0]*9                                                     

			# Counting total count of values in different octs
			for i in octant:                                                
				if(i==1):
					col_list[1]=col_list[1]+1
				elif(i==-1):
					col_list[2]=col_list[2]+1
				elif(i==2):
					col_list[3]=col_list[3]+1
				elif(i==-2):
					col_list[4]=col_list[4]+1
				elif(i==3):
					col_list[5]=col_list[5]+1
				elif(i==-3):
					col_list[6]=col_list[6]+1
				elif(i==4):
					col_list[7]=col_list[7]+1
				elif(i==-4):
					col_list[8]=col_list[8]+1

			CodeYelow = "00FFFF00"


			# Creating overall count row
			col_list[0]='Overall Count'                                        
			oct_range_matrix.append(col_list)          


			 # Writing overall count in worksheet                                 
			for i in range(13,22):                                         
				sheet.cell(row=4,column=i+1).value=col_list[i-13]
				sheet.cell(row=4,column=i+1).border=bor
			# Removing the header from list	
			col_list.pop(0)                     


			# Computing the rank list                                
			rankk=rankListFind(col_list)             

			 # Finding the rank 1 octant and appending in rank1_list                      
			rank1_list.append(rank_1(rankk))                         
			mtrix_rank.append(rankk)            

			# Writing overall count in worksheet                           
			for i in range(8):                                              
				sheet.cell(row=4,column=23+i).value=mtrix_rank[0][i]
				sheet.cell(row=4,column=23+i).border=bor
				if(mtrix_rank[0][i]==1):
					sheet.cell(row=4,column=23+i).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
			sheet.cell(row=4,column=31).value=rank1_list[0]
			sheet.cell(row=4,column=32).value=octant_name_id_mapping[str(rank1_list[0])]
			sheet.cell(row=4,column=31).border=bor
			sheet.cell(row=4,column=32).border=bor
									
			
			# Finding the number of points given in the input
			m=len(octant)

			# Resetting the values in the list 'count'                                              
			col_list=[0]*9          

			# To keep track of the current index.                                           
			b=0                     

			# To keep track of row in worksheet                                        
			j=4               

			 # Counting number of values in different octs in mod range                                              
			for i in octant:                                               
				if(i==1):
					col_list[1]=col_list[1]+1
				elif(i==-1):
					col_list[2]=col_list[2]+1
				elif(i==2):
					col_list[3]=col_list[3]+1
				elif(i==-2):
					col_list[4]=col_list[4]+1
				elif(i==3):
					col_list[5]=col_list[5]+1
				elif(i==-3):
					col_list[6]=col_list[6]+1
				elif(i==4):
					col_list[7]=col_list[7]+1
				elif(i==-4):
					col_list[8]=col_list[8]+1
				b=b+1                                                     
				if(b%mod==1):                                              
					col_list[0]=str(b-1)+'-'                       
				elif(b%mod==0 or b==m):
					col_list[0]=col_list[0]+str(b-1)    


					# Adding the columns of rank, rank1 and octant_name in the worksheet
					for i in range(13,22):                                
						sheet.cell(row=j+1,column=i+1).value=col_list[i-13]
						sheet.cell(row=j+1,column=i+1).border=bor
					col_list.pop(0)                                         
					rankk=rankListFind(col_list)                           
					rank1_list.append(rank_1(rankk))                
					mtrix_rank.append(rankk)                            
					

					# Adding columns of rank, rank1 and octant_name in the worksheet
					for i in range(8):                                              
						sheet.cell(row=j+1,column=23+i).value=mtrix_rank[j-3][i]
						sheet.cell(row=j+1,column=23+i).border=bor
						if(mtrix_rank[j-3][i]==1):
							sheet.cell(row=j+1,column=23+i).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
					
					sheet.cell(row=j+1,column=31).value=rank1_list[j-3]
					sheet.cell(row=j+1,column=32).value=octant_name_id_mapping[str(rank1_list[j-3])]
					sheet.cell(row=j+1,column=31).border=bor
					sheet.cell(row=j+1,column=32).border=bor
					
					j=j+1
					oct_range_matrix.append(col_list)
					col_list=[0]*9                                
													
			rank1_list.pop(0)                                         
			j+=4
			sheet.cell(row=j,column=29).value='Octant ID'        
			sheet.cell(row=j,column=30).value='Octant Name'
			sheet.cell(row=j,column=31).value='Count of Rank 1 Mod Values'
			sheet.cell(row=j,column=29).border=bor                 
			sheet.cell(row=j,column=30).border=bor
			sheet.cell(row=j,column=31).border=bor
			

			# Appending the table of count of rank1 mod values
			for i in range(8):                              
				sheet.cell(row=j+1+i,column=29).value=opp_dict_mapping[i]
				sheet.cell(row=j+1+i,column=30).value=octant_name_id_mapping[str(opp_dict_mapping[i])]
				sheet.cell(row=j+1+i,column=31).value=count_rank_1(rank1_list,opp_dict_mapping[i])
				sheet.cell(row=j+1+i,column=29).border=bor
				sheet.cell(row=j+1+i,column=30).border=bor
				sheet.cell(row=j+1+i,column=31).border=bor
		
		def oct_longest_subsequence_count_within_range():
			
			# Header list
			r=['Count','Longest Subsequence Length','Count']

			# Writing header of table to worksheet        
			for i in range(3):                                     
				sheet.cell(row=3,column=45+i).value=r[i] 
				sheet.cell(row=3,column=45+i).border=bor    


			octs=[]

			# Writing octants on leftmost column of the table
			for i in range(2,10,2):                                    
				sheet.cell(row=i+2,column=45).value=i//2
				octs.append(i//2)
				sheet.cell(row=i+3,column=45).value=-(i//2) 
				octs.append(-i//2)
				sheet.cell(row=i+2,column=45).border=bor
				sheet.cell(row=i+3,column=45).border=bor                                
			
			# Dictionary for mapping 
			dict_mapping={}                                              
			for i in range(0,4):                                            
				dict_mapping[i+1]=2*i+1-1
				dict_mapping[-(i+1)]=2*(i+1)-1

			# List for storing number of long subsequence		
			col_list=[0]*8                                             
			longest_length=[0]*8                                    
			back=octant[0]

			# Length of current oct
			l=1                                                           
			n=len(octant)

			 # Temporary variable to store range
			max_temp=[0]                                                  
			ranges= [[] for x in repeat(None, 8)]            

			 # Finding the length and number of longest subsequence
			for i in range(1,n+1):  
				  # Processing the whole if last is reached                
				if(i==n):                                        
					if(longest_length[dict_mapping[back]]<l):                        
						longest_length[dict_mapping[back]]=l
						col_list[dict_mapping[back]]=1

						# Writing ending range in temp
						max_temp.append(df['T'][i-1])            
						ranges[dict_mapping[back]].clear()         
						ranges[dict_mapping[back]].append(max_temp)           
					elif(longest_length[dict_mapping[back]]==l):
						col_list[dict_mapping[back]]+=1
						max_temp.append(df['T'][i-1])
						ranges[dict_mapping[back]].append(max_temp) 
				 # If prev and current values are same, increase current length by 1		               
				elif(back==octant[i]):                                      
					l+=1

				# Else process the previous octant values and start with new octant
				else:                                                      
					if(longest_length[dict_mapping[back]]<l):
						longest_length[dict_mapping[back]]=l
						col_list[dict_mapping[back]]=1
						ranges[dict_mapping[back]].clear()               
						max_temp.append(df['T'][i-1])        
						ranges[dict_mapping[back]].append(max_temp)                 
					elif(longest_length[dict_mapping[back]]==l):
						col_list[dict_mapping[back]]+=1
						max_temp.append(df['T'][i-1])
						ranges[dict_mapping[back]].append(max_temp)             
					max_temp=[df['T'][i]]                     
					l=1
					back=octant[i]                            

			# Writing the number and length of longest subsequence in table
			for i in range(2,10):                                    
				sheet.cell(row=i+2,column=46).value=longest_length[i-2]
				sheet.cell(row=i+2,column=47).value=col_list[i-2]
				sheet.cell(row=i+2,column=46).border=bor
				sheet.cell(row=i+2,column=47).border=bor
			b=2                                                        
			sheet.cell(row=b+1,column=49).value='Octant ###'                  
			sheet.cell(row=b+1,column=50).value='Longest Subsequence Length'
			sheet.cell(row=b+1,column=51).value='Count'
			sheet.cell(row=b+1,column=49).border=bor
			sheet.cell(row=b+1,column=50).border=bor
			sheet.cell(row=b+1,column=51).border=bor
			
			b+=2
			for i in range(8):
				sheet.cell(row=b,column=49).value=octs[i]            
				sheet.cell(row=b,column=50).value=longest_length[i]
				sheet.cell(row=b,column=51).value=col_list[i]
				sheet.cell(row=b+1,column=49).value='Time'             
				sheet.cell(row=b+1,column=50).value='From'
				sheet.cell(row=b+1,column=51).value='To'
				sheet.cell(row=b,column=49).border=bor      
				sheet.cell(row=b,column=50).border=bor
				sheet.cell(row=b,column=51).border=bor
				sheet.cell(row=b+1,column=49).border=bor
				sheet.cell(row=b+1,column=50).border=bor
				sheet.cell(row=b+1,column=51).border=bor
				x=ranges[i]
				b+=2
				for j in x:
					# Writing ranges in worksheet
					sheet.cell(row=b,column=50).value=j[0]     
					sheet.cell(row=b,column=51).value=j[1]
					sheet.cell(row=b,column=49).border=bor  
					sheet.cell(row=b,column=50).border=bor
					sheet.cell(row=b,column=51).border=bor
					b+=1
			
		def octant_transition_count(mod=5000):
			j=1
			n=len(octant)
			sheet.cell(row=j,column=35).value='Overall Transition Count' 
			sheet.cell(row=j+3,column=34).value='From'
			sheet.cell(row=j+1,column=36).value='To'
			j+=2
			
			oct_range_matrix = [ [0]*9 for i in range(9)]                    
			
			# Appending header row and header column in the matrix
			for i in range(0,4):                                        
				oct_range_matrix[0][2*i+1]=(i+1)
				oct_range_matrix[0][2*(i+1)]=-(i+1)
			for i in range(0,9):
				oct_range_matrix[i][0]=oct_range_matrix[0][i]
			oct_range_matrix[0][0]='Octant #'

			 # Creating dictionary for mapping 
			dict_mapping={}                                             
			for i in range(0,4):
				dict_mapping[i+1]=2*i+1
				dict_mapping[-(i+1)]=2*(i+1)

			  # Finding row and column of matrix from transition values
			def find_row_col(x,y):                     
				head_comp=[dict_mapping[x],dict_mapping[y]]
				return head_comp
			
			def find_max_ele(head_comp):
				max_temp=head_comp.copy()
				max_temp.pop(0)
				large=0
				for i in max_temp:
					if(large<i):
						large=i
				return large

			back=octant[0]

			# Filling overall transition matrix                                              
			for i in range(1,n):                       
				head_comp=find_row_col(back,octant[i])                        
				oct_range_matrix[head_comp[0]][head_comp[1]]+=1
				back=octant[i]

			CodeYelow = "00FFFF00"

			# Writing the overall transition matrix in worksheet
			for i in range(9):                                  
				temp_head_comp=oct_range_matrix[i]
				large=find_max_ele(temp_head_comp)
				for b in range(13,22):
					sheet.cell(row=j+i,column=b+22).value=oct_range_matrix[i][b-13]
					sheet.cell(row=j+i,column=b+22).border=bor
					if(i>0 and oct_range_matrix[i][b-13]==large):
						sheet.cell(row=j+i,column=b+22).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
					if(i!=0 and b!=13):
						oct_range_matrix[i][b-13]=0

			# temp-> No. of mod transition tables	
			max_temp=n//mod+1                                              
			j+=1

			# One iteration for each mod transition table
			for t in range(max_temp):                                   
				j+=11
				nam1=''

				# Writing Table name in worksheet
				sheet.cell(row=j,column=35).value='Mod Transition Count'     
				sheet.cell(row=j+3,column=34).value='From'
				sheet.cell(row=j+1,column=36).value='To'
				nam1=str(t*mod)+'-'
				if((t+1)*mod-1>n-1):
					nam1+=str(n-1)
				else:
					nam1+=str((t+1)*mod-1)   
				sheet.cell(row=j+1,column=35).value=nam1
				j+=2

				# Incrementing matrix cell corresponding to transition values
				for i in range(t*mod,min(n-1,(t+1)*mod)):                 
					head_comp=find_row_col(octant[i],octant[i+1])
					oct_range_matrix[head_comp[0]][head_comp[1]]+=1

				# Writing the transition mod matrix in worksheet
				for i in range(0,9):                                     
					temp_head_comp=oct_range_matrix[i]
					if(i>0):
						large=find_max_ele(temp_head_comp)
					for b in range(13,22):
						sheet.cell(row=j+i,column=b+22).value=oct_range_matrix[i][b-13]
						sheet.cell(row=j+i,column=b+22).border=bor
						if(i>0 and oct_range_matrix[i][b-13]==large):
							sheet.cell(row=j+i,column=b+22).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
						if(i!=0 and b!=13):

							# Resetting matrix for next mod iteration
							oct_range_matrix[i][b-13]=0                           


		octant_transition_count(mod)
		range_name_oct(mod)
		oct_longest_subsequence_count_within_range()
		s=s.name[:-5]
		path_out='output/'+s+' cm_vel_octant_analysis_mod_'+str(mod)+'.xlsx'
		wb.save(path_out)

		data_1 = BytesIO()
		wb.save(data_1)
		filename = s+ "cm_vel_octant_analysis_mod_"+str(val)+".xlsx" 
		st.text(filename) 
		st.download_button(label="Download File", file_name=s+
                           "cm_vel_octant_analysis_mod_"+str(val)+".xlsx", data=data_1)
	# a = pathlib.Path(file).parent.parent.resolve()
	# input_files = os.path.join(a, "proj2\input")
	# # input_files=os.listdir('input')
	# for i in range(len(input_files)):
	# 	workingFile(mod,file)
	wf(file,mod)



# **MULTIPLE file****

def octant_analysis_mutiple(mod,path):
	from pandas import read_excel
	import openpyxl
	from openpyxl import workbook,load_workbook
	from openpyxl import Workbook
	from openpyxl.styles import PatternFill
	from itertools import repeat

	def workingFile(s,file,mod):							
		inputFilePath=s									
		df=read_excel(inputFilePath)									
		wb=Workbook()											
		sheet=wb.active													
		
		for i in range(11):										
			sheet.cell(row=2,column=i+1).value=h_c[i]	
		octant=[]
		# sheet.cell(2,1).value = 'T'
		# sheet.cell(2,2).value = 'U'
		# sheet.cell(2,3).value = 'V'
		# sheet.cell(2,4).value = 'W'
		# sheet.cell(2,5).value = 'U Avg'
		# sheet.cell(2,6).value = 'V Avg'
		# sheet.cell(2,7).value = 'W Avg'
		# sheet.cell(2,8).value = "U'=U-U avg"
		# sheet.cell(2,9).value = "V'=V-V avg"
		# sheet.cell(2,10).value = "W'=W-W avg"
		# sheet.cell(2,11).value = 'Octant'

		u_avg=df['U'].mean()                                                 
		v_avg=df['V'].mean()
		w_avg=df['W'].mean()
		sheet.cell(row=1,column=14).value='Overall Octant Count'
		sheet.cell(row=1,column=45).value='Longest Subsequence Length'
		sheet.cell(row=1,column=49).value='Longest Subsquence Length with Range'
		sheet['E3'] = u_avg
		sheet['F3'] = v_avg
		sheet['G3'] = w_avg


		for i in df.index:
			sheet.cell(row=i+3,column=1).value=df['T'][i]
			sheet.cell(row=i+3,column=2).value=df['U'][i]
			sheet.cell(row=i+3,column=3).value=df['V'][i]
			sheet.cell(row=i+3,column=4).value=df['W'][i]
			sheet.cell(row=i+3,column=8).value=round(df['U'][i]-u_avg,3)
			sheet.cell(row=i+3,column=9).value=round(df['V'][i]-v_avg,3)
			sheet.cell(row=i+3,column=10).value=round(df['W'][i]-w_avg,3)
			sheet.cell(row=i+3,column=11).value=oct_find(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg)
			octant.append(oct_find(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg))
		
		def range_name_oct(mod=5000):
			octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}

			# creating dictionary for mapping 
			dict_mapping={}                            

			# Creating dictionary with opposite key value pair than 'dict_mapping'                              
			opp_dict_mapping={}                                                      
			
			for i in range(4):                                            
				dict_mapping[i+1]=2*i+1-1                                           
				dict_mapping[-(i+1)]=2*(i+1)-1
				opp_dict_mapping[2*i+1-1]=i+1
				opp_dict_mapping[2*(i+1)-1]=-(i+1)

			# this function is used to find rank of each octant
			def rankListFind(h_c):                                     
				temp_head_comp=h_c.copy()
				temp_head_comp.sort(reverse=True)
				res=[]

				for i in h_c:
					for j in range(8):
						if(i==temp_head_comp[j]):
							res.append(j+1)
							break
				# Returning the ranked list		
				return res                                                  
			
			# The given list is for finding the rank 1 octants
			def find_1st_rank(h_c):                                         
				for i in range(8):
					if(h_c[i]==1):
						return opp_dict_mapping[i]

			# Finding the count of rank 1 in the rank 1 mod values of octant x
			def Rank1Count(h_c,x):                                         
				ct=0
				for i in h_c:
					if(x==i):
						ct+=1

				# Return the count		
				return ct                                                  
			
			 # Storing rank list for different mod values in matrix
			mtrix_rank=[] 

			# storing the octs having rank 1 in diff. mod ranges and overall                                   
			rank1_list=[] 
			try:                                                  
				sheet=wb.active
			except:
				print("sheet activation error")	

			# Adding string 'user input' at specified place
			sheet['M4']='Mod '+str(mod)                                           
			# For storing oct within ranges as 2-d matrix.
			oct_range_matrix=[]   
			# Storing elements of 9 columns in a list                                                       
			col_list=[0]*9                                                     

			 # Storing header list in 'col_list' list
			col_list[0]='Octant ID'                                           

			for i in range(0,4):
				col_list[2*i+1]=(i+1)
				col_list[2*(i+1)]=-(i+1)

			# Appending header list in matrix	
			oct_range_matrix.append(col_list)          

			# Appending header list in worksheet                                  
			for i in range(13,22):                                          
				sheet.cell(row=3,column=i+1).value=col_list[i-13]
				sheet.cell(row=3,column=i+1).border=bor
				if(i>13):
					sheet.cell(row=3,column=i+9).value='Rank Octant '+str(col_list[i-13])
					sheet.cell(row=3,column=i+9).border=bor
			sheet.cell(row=3,column=31).value='Rank1 Octant ID'
			sheet.cell(row=3,column=32).value='Rank1 Octant Name'
			sheet.cell(row=3,column=31).border=bor
			sheet.cell(row=3,column=32).border=bor

			# Changing values to default in list 'col_list'
			col_list=[0]*9                                                     

			# Counting total count of values in different octs
			for i in octant:                                                
				if(i==1):
					col_list[1]=col_list[1]+1
				elif(i==-1):
					col_list[2]=col_list[2]+1
				elif(i==2):
					col_list[3]=col_list[3]+1
				elif(i==-2):
					col_list[4]=col_list[4]+1
				elif(i==3):
					col_list[5]=col_list[5]+1
				elif(i==-3):
					col_list[6]=col_list[6]+1
				elif(i==4):
					col_list[7]=col_list[7]+1
				elif(i==-4):
					col_list[8]=col_list[8]+1

			CodeYelow = "00FFFF00"

			# Creating overall count row
			col_list[0]='Overall Count'                                        
			oct_range_matrix.append(col_list)          


			 # Appending overall count in worksheet                                 
			for i in range(13,22):                                         
				sheet.cell(row=4,column=i+1).value=col_list[i-13]
				sheet.cell(row=4,column=i+1).border=bor
			# Removing the header from list	
			col_list.pop(0)                     


			# Computing the rank list                                
			rankk=rankListFind(col_list)             

			 # Finding the rank 1 octant and appending in rank1_list                      
			rank1_list.append(find_1st_rank(rankk))                         
			mtrix_rank.append(rankk)            

			# Appending overall count in worksheet                           
			for i in range(8):                                              
				sheet.cell(row=4,column=23+i).value=mtrix_rank[0][i]
				sheet.cell(row=4,column=23+i).border=bor
				if(mtrix_rank[0][i]==1):
					sheet.cell(row=4,column=23+i).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
			sheet.cell(row=4,column=31).value=rank1_list[0]
			sheet.cell(row=4,column=32).value=octant_name_id_mapping[str(rank1_list[0])]
			sheet.cell(row=4,column=31).border=bor
			sheet.cell(row=4,column=32).border=bor
									
			
			# Calculating the number of points given in the input
			m=len(octant)     

			# Changing the values to default in the list 'count'                                              
			col_list=[0]*9          

			# To keep track of the current index. 
			b=0                     

			# To keep track of row in worksheet                                         
			j=4               

			 # Counting number of values in different octs in mod range                                              
			for i in octant:                                               
				if(i==1):
					col_list[1]=col_list[1]+1
				elif(i==-1):
					col_list[2]=col_list[2]+1
				elif(i==2):
					col_list[3]=col_list[3]+1
				elif(i==-2):
					col_list[4]=col_list[4]+1
				elif(i==3):
					col_list[5]=col_list[5]+1
				elif(i==-3):
					col_list[6]=col_list[6]+1
				elif(i==4):
					col_list[7]=col_list[7]+1
				elif(i==-4):
					col_list[8]=col_list[8]+1
				b=b+1                                                     
				if(b%mod==1):                                              
					col_list[0]=str(b-1)+'-'                       
				elif(b%mod==0 or b==m):
					col_list[0]=col_list[0]+str(b-1)    

					# Adding the columns of rank, rank1 and octant_name in the worksheet
					for i in range(13,22):                                
						sheet.cell(row=j+1,column=i+1).value=col_list[i-13]
						sheet.cell(row=j+1,column=i+1).border=bor
					col_list.pop(0)                                         
					rankk=rankListFind(col_list)                           
					rank1_list.append(find_1st_rank(rankk))                
					mtrix_rank.append(rankk)                            
					
					# Adding columns of rank, rank1 and octant_name in the worksheet
					for i in range(8):                                              
						sheet.cell(row=j+1,column=23+i).value=mtrix_rank[j-3][i]
						sheet.cell(row=j+1,column=23+i).border=bor
						if(mtrix_rank[j-3][i]==1):
							sheet.cell(row=j+1,column=23+i).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
					
					sheet.cell(row=j+1,column=31).value=rank1_list[j-3]
					sheet.cell(row=j+1,column=32).value=octant_name_id_mapping[str(rank1_list[j-3])]
					sheet.cell(row=j+1,column=31).border=bor
					sheet.cell(row=j+1,column=32).border=bor
					
					j=j+1
					oct_range_matrix.append(col_list)
					col_list=[0]*9                                
													
			rank1_list.pop(0)                                         
			j+=4
			sheet.cell(row=j,column=29).value='Octant ID'        
			sheet.cell(row=j,column=30).value='Octant Name'
			sheet.cell(row=j,column=31).value='Count of Rank 1 Mod Values'
			sheet.cell(row=j,column=29).border=bor                 
			sheet.cell(row=j,column=30).border=bor
			sheet.cell(row=j,column=31).border=bor
			
			# Appending the table of count of rank1 mod values
			for i in range(8):                              
				sheet.cell(row=j+1+i,column=29).value=opp_dict_mapping[i]
				sheet.cell(row=j+1+i,column=30).value=octant_name_id_mapping[str(opp_dict_mapping[i])]
				sheet.cell(row=j+1+i,column=31).value=Rank1Count(rank1_list,opp_dict_mapping[i])
				sheet.cell(row=j+1+i,column=29).border=bor
				sheet.cell(row=j+1+i,column=30).border=bor
				sheet.cell(row=j+1+i,column=31).border=bor
		
		def oct_long_subseq_count_with_rang():
			
			# Header list
			r=['Count','Longest Subsequence Length','Count']

			# Appending the header of table to worksheet        
			for i in range(3):                                     
				sheet.cell(row=3,column=45+i).value=r[i] 
				sheet.cell(row=3,column=45+i).border=bor    


			octs=[]

			# Appending octants on leftmost column of the table
			for i in range(2,10,2):                                    
				sheet.cell(row=i+2,column=45).value=i//2
				octs.append(i//2)
				sheet.cell(row=i+3,column=45).value=-(i//2) 
				octs.append(-i//2)
				sheet.cell(row=i+2,column=45).border=bor
				sheet.cell(row=i+3,column=45).border=bor                                
			
			# Dictionary for mapping 
			dict_mapping={}                                              
			for i in range(0,4):                                            
				dict_mapping[i+1]=2*i+1-1
				dict_mapping[-(i+1)]=2*(i+1)-1

			# List for storing number of long subsequence
			col_list=[0]*8                                             
			longest_length=[0]*8                                    
			back=octant[0]

			# Length of current oct
			l=1                                                           
			n=len(octant)

			 # Storing range in temporary variable....
			max_temp=[0]                                                  
			ranges= [[] for x in repeat(None, 8)]            

			 # Finding the length and number of longest subsequence
			for i in range(1,n+1):  
				  # Processing the whole if last is reached                
				if(i==n):                                        
					if(longest_length[dict_mapping[back]]<l):                        
						longest_length[dict_mapping[back]]=l
						col_list[dict_mapping[back]]=1

						# Adding ending range in temp....
						max_temp.append(df['T'][i-1])            
						ranges[dict_mapping[back]].clear()         
						ranges[dict_mapping[back]].append(max_temp)           
					elif(longest_length[dict_mapping[back]]==l):
						col_list[dict_mapping[back]]+=1
						max_temp.append(df['T'][i-1])
						ranges[dict_mapping[back]].append(max_temp) 
				 # Increasing the current length by 1 if prev and current values are same....		               
				elif(back==octant[i]):                                      
					l+=1

				# Else process the previous octant values and start with new octant
				else:                                                      
					if(longest_length[dict_mapping[back]]<l):
						longest_length[dict_mapping[back]]=l
						col_list[dict_mapping[back]]=1
						ranges[dict_mapping[back]].clear()               
						max_temp.append(df['T'][i-1])        
						ranges[dict_mapping[back]].append(max_temp)                 
					elif(longest_length[dict_mapping[back]]==l):
						col_list[dict_mapping[back]]+=1
						max_temp.append(df['T'][i-1])
						ranges[dict_mapping[back]].append(max_temp)             
					max_temp=[df['T'][i]]                     
					l=1
					back=octant[i]                            

			# Appending the length and number of longest subsequence in table
			for i in range(2,10):                                    
				sheet.cell(row=i+2,column=46).value=longest_length[i-2]
				sheet.cell(row=i+2,column=47).value=col_list[i-2]
				sheet.cell(row=i+2,column=46).border=bor
				sheet.cell(row=i+2,column=47).border=bor
			b=2                                                        
			sheet.cell(row=b+1,column=49).value='Octant ###'                  
			sheet.cell(row=b+1,column=50).value='Longest Subsequence Length'
			sheet.cell(row=b+1,column=51).value='Count'
			sheet.cell(row=b+1,column=49).border=bor
			sheet.cell(row=b+1,column=50).border=bor
			sheet.cell(row=b+1,column=51).border=bor
			
			b+=2
			for i in range(8):
				sheet.cell(row=b,column=49).value=octs[i]            
				sheet.cell(row=b,column=50).value=longest_length[i]
				sheet.cell(row=b,column=51).value=col_list[i]
				sheet.cell(row=b+1,column=49).value='Time'             
				sheet.cell(row=b+1,column=50).value='From'
				sheet.cell(row=b+1,column=51).value='To'
				sheet.cell(row=b,column=49).border=bor      
				sheet.cell(row=b,column=50).border=bor
				sheet.cell(row=b,column=51).border=bor
				sheet.cell(row=b+1,column=49).border=bor
				sheet.cell(row=b+1,column=50).border=bor
				sheet.cell(row=b+1,column=51).border=bor
				x=ranges[i]
				b+=2
				for j in x:
					# Writing ranges in worksheet
					sheet.cell(row=b,column=50).value=j[0]     
					sheet.cell(row=b,column=51).value=j[1]
					sheet.cell(row=b,column=49).border=bor  
					sheet.cell(row=b,column=50).border=bor
					sheet.cell(row=b,column=51).border=bor
					b+=1
			
