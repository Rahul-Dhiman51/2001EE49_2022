from datetime import datetime
start_time = datetime.now()

def octant_range_names(mod=5000):
 from openpyxl import load_workbook
 import os
 import numpy as np
 from operator import itemgetter

 os.system('cls')

 d1 = []    # declaring new list
 b1 = []    # declaring new list
 a1 = []    # declaring new list
 c1 = []    # declaring new list
 b2 = []    # declaring list for difference 
 a2 = []    # declaring list for difference
 c2 = []    # declaring list for difference
 octants = []   # declaring list for octant storing

 work_book=load_workbook("octant_input.xlsx")
 spreadsheet=work_book["Sheet1"] 
 num_1=spreadsheet.max_row
 for i in range(2,num_1+1):
  d1.append(spreadsheet.cell(row=i,column=1).value)
  a1.append(spreadsheet.cell(row=i,column=2).value)
  b1.append(spreadsheet.cell(row=i,column=3).value)
  c1.append(spreadsheet.cell(row=i,column=4).value)

 b_mean_value=np.mean(b1, dtype=np.float64) # for mean of b1
 a_mean_value=np.mean(a1, dtype=np.float64) # for mean of a1
 c_mean_value=np.mean(c1, dtype=np.float64) # for mean of c1
 
 for i in range(2,num_1+1):
  a2.append(float(spreadsheet.cell(row=i,column=2).value)-a_mean_value) # pushing the differences 
  b2.append(float(spreadsheet.cell(row=i,column=3).value)-b_mean_value) # pushing the differences
  c2.append(float(spreadsheet.cell(row=i,column=4).value)-c_mean_value)	# pushing the differences  

  # Finding octant with the help of for loop
 for i in range(0,num_1-1):
  if ((a2[i]>=0) and (b2[i]>=0) and (c2[i]>=0 )):
   octants.append(1)
  elif((a2[i]>=0) and (b2[i]>=0) and (c2[i]<0 )):
   octants.append(-1)
  elif((a2[i]<0) and (b2[i]>=0) and (c2[i]>=0 )):
   octants.append(2)
  elif((a2[i]<0) and (b2[i]>=0) and (c2[i]<0 )):
   octants.append(-2)
  elif((a2[i]>=0) and (b2[i]<0) and (c2[i]>=0 )):
   octants.append(4)
  elif((a2[i]>=0) and (b2[i]<0) and (c2[i]<0 )):
   octants.append(-4)    
  elif((a2[i]<0) and (b2[i]<0) and (c2[i]>=0 )):
   octants.append(3)
  else:
   octants.append(-3) 
     
  ct1=0         # counting of octant 1
  ct2=0         # counting of octant -1
  ct3=0         # counting of octant 2
  ct4=0         # counting of octant -2
  ct5=0         # counting of octant 3
  ct6=0         # counting of octant -3
  ct7=0         # counting of octant 4
  ct8=0         # counting of octant -4

  #finding the octants using loop
 for i in range(0,num_1-1):
  if (octants[i]==1):
   ct1=ct1+1
  elif(octants[i]==-1):
   ct2=ct2+1
  elif(octants[i]==2):
   ct3=ct3+1
  elif(octants[i]==-2):
   ct4=ct4+1
  elif(octants[i]==3):
   ct5=ct5+1
  elif(octants[i]==-3):
   ct6=ct6+1    
  elif(octants[i]==4):
   ct7=ct7+1
  else:
   ct8=ct8+1

 ranks=[]
 rank1_idname=[]
 reqrd=[]
 reqrd.append([ct1,1])
 reqrd.append([ct2,2])
 reqrd.append([ct3,3])
 reqrd.append([ct4,4])
 reqrd.append([ct5,5])
 reqrd.append([ct6,6])
 reqrd.append([ct7,7])
 reqrd.append([ct8,8]) 

 reqrd.sort()

 reqrd_ranks=[0,0,0,0,0,0,0,0]
 reqrd_ranks[reqrd[0][1]-1]=8
 reqrd_ranks[reqrd[1][1]-1]=7
 reqrd_ranks[reqrd[2][1]-1]=6
 reqrd_ranks[reqrd[3][1]-1]=5
 reqrd_ranks[reqrd[4][1]-1]=4
 reqrd_ranks[reqrd[5][1]-1]=3
 reqrd_ranks[reqrd[6][1]-1]=2
 reqrd_ranks[reqrd[7][1]-1]=1
 ranks.append(reqrd_ranks)

 rank_idname=[]

 if reqrd_ranks[0]==1:
  rank_idname=[1,"Internal outward Interaction"]
  rank1_idname.append(rank_idname)
 elif reqrd_ranks[1]==1:
  rank_idname=[-1,"External outward Interaction"]
  rank1_idname.append(rank_idname)
 elif reqrd_ranks[2]==1:
  rank_idname=[2,"External Ejection"]
  rank1_idname.append(rank_idname) 
 elif reqrd_ranks[3]==1:
  rank_idname=[-2,"Internal Ejection"]
  rank1_idname.append(rank_idname) 
 elif reqrd_ranks[4]==1:
  rank_idname=[3,"External inward Interaction"]
  rank1_idname.append(rank_idname) 
 elif reqrd_ranks[5]==1:
  rank_idname=[-3,"Internal inward Interaction"]
  rank1_idname.append(rank_idname) 
 elif reqrd_ranks[6]==1:
  rank_idname=[4,"Internal Sweep"]
  rank1_idname.append(rank_idname) 
 elif reqrd_ranks[7]==1:
  rank_idname=[-4,"External Sweep"]
  rank1_idname.append(rank_idname) 

 # Making lists of different octants in mod range

 octants_ct1=[]      
 octants_ct2=[] 
 octants_ct3=[]     
 octants_ct4=[]
 octants_ct5=[]
 octants_ct6=[] 
 octants_ct7=[]
 octants_ct8=[]

  #count of octants in mod range...

 oct_1=0 
 oct_2=0
 oct_3=0
 oct_4=0
 oct_5=0
 oct_6=0
 oct_7=0
 oct_8=0

 k= int((num_1-2)/mod) +1     # k gives the value of interval.
 sn_1=0
 sn_2=0
 sn_3=0
 sn_4=0
 sn_5=0
 sn_6=0
 sn_7=0
 sn_8=0
  
 for i in range(k):
  y=mod*i
  for j in range(mod):	
   if(j+y<num_1-1):	                                    
    if (octants[j+y]==1):
     oct_1=oct_1+1 
    elif(octants[j+y]==-1):
     oct_2=oct_2+1
    elif(octants[j+y]==2):
     oct_3=oct_3+1
    elif(octants[j+y]==-2):
     oct_4=oct_4+1
    elif(octants[j+y]==3):
     oct_5=oct_5+1
    elif(octants[j+y]==-3):
     oct_6=oct_6+1    
    elif(octants[j+y]==4):
     oct_7=oct_7+1
    elif(octants[j+y]==-4):
     oct_8=oct_8+1

  octants_ct1.append(oct_1)
  octants_ct2.append(oct_2) 
  octants_ct3.append(oct_3)
  octants_ct4.append(oct_4)
  octants_ct5.append(oct_5)
  octants_ct6.append(oct_6)
  octants_ct7.append(oct_7)
  octants_ct8.append(oct_8)

  reqrd=[]
  reqrd.append([oct_1,1])
  reqrd.append([oct_2,2])
  reqrd.append([oct_3,3])
  reqrd.append([oct_4,4])
  reqrd.append([oct_5,5])
  reqrd.append([oct_6,6])
  reqrd.append([oct_7,7])
  reqrd.append([oct_8,8])

  reqrd.sort()
  reqrd_ranks=[0,0,0,0,0,0,0,0]
  reqrd_ranks[reqrd[0][1]-1]=8
  reqrd_ranks[reqrd[1][1]-1]=7
  reqrd_ranks[reqrd[2][1]-1]=6
  reqrd_ranks[reqrd[3][1]-1]=5
  reqrd_ranks[reqrd[4][1]-1]=4
  reqrd_ranks[reqrd[5][1]-1]=3
  reqrd_ranks[reqrd[6][1]-1]=2
  reqrd_ranks[reqrd[7][1]-1]=1

  ranks.append(reqrd_ranks)
  rank_idname=[]
  rank_idname=[]
  if reqrd_ranks[0]==1:
   rank_idname=[1,"Internal outward Interaction"]
   rank1_idname.append(rank_idname)
   sn_1=sn_1+1
  elif reqrd_ranks[1]==1:
   rank_idname=[-1,"External outward Interaction"]
   rank1_idname.append(rank_idname)
   sn_2=sn_2+1
  elif reqrd_ranks[2]==1:
   rank_idname=[2,"External Ejection"]
   rank1_idname.append(rank_idname)
   sn_3=sn_3+1 
  elif reqrd_ranks[3]==1:
   rank_idname=[-2,"Internal Ejection"]
   rank1_idname.append(rank_idname)
   sn_4=sn_4+1 
  elif reqrd_ranks[4]==1:
   rank_idname=[3,"External inward Interaction"]
   rank1_idname.append(rank_idname)
   sn_5=sn_5+1 
  elif reqrd_ranks[5]==1:
   rank_idname=[-3,"Internal inward Interaction"]
   rank1_idname.append(rank_idname)
   sn_6=sn_6+1 
  elif reqrd_ranks[6]==1:
   rank_idname=[4,"Internal Sweep"]
   rank1_idname.append(rank_idname)
   sn_7=sn_7+1 
  elif reqrd_ranks[7]==1:
   rank_idname=[-4,"External Sweep"]
   rank1_idname.append(rank_idname)
   sn_8=sn_8+1 
 
  oct_1=0
  oct_2=0
  oct_3=0
  oct_4=0
  oct_5=0
  oct_6=0
  oct_7=0
  oct_8=0
  
 from openpyxl import Workbook 
 book=Workbook()
 spreadsheet= book.active    
 rows=[] 
 rows.append(["Time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant","","OctantID","+1","-1","+2","-2","+3","-3","+4","-4","Rank of 1","Rank of -1","Rank of 2","Rank of -2","Rank of 3","Rank of -3","Rank of 4","Rank of -4","Rank1 Octant ID","Rank1 Octant Name"])
 j=0
 a=0
 b=0
 for x in range(num_1-2):
  if(x==0):
   rows.append([d1[x],a1[x],b1[x],c1[x],a_mean_value,b_mean_value,c_mean_value,a2[x],b2[x],c2[x],octants[x],"","Overall count",ct1,ct2,ct3,ct4,ct5,ct6,ct7,ct8,ranks[x][0],ranks[x][1],ranks[x][2],ranks[x][3],ranks[x][4],ranks[x][5],ranks[x][6],ranks[x][7],rank1_idname[x][0],rank1_idname[x][1]])
  elif(x==1):
   y="mod "+str(mod)		
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"User input",y,"","","","","","","",""])
  elif(x>=2 and x<2+k):
   if(x==1+k):# it will work only if x=1+k
    z=j*mod 	
    y=(j+1)*mod
    y=str(z)+"-"+str(num_1-2)		 
    rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"",y,octants_ct1[x-2],octants_ct2[x-2],octants_ct3[x-2],octants_ct4[x-2],octants_ct5[x-2],octants_ct6[x-2],octants_ct7[x-2],octants_ct8[x-2],ranks[x-1][0],ranks[x-1][1],ranks[x-1][2],ranks[x-1][3],ranks[x-1][4],ranks[x-1][5],ranks[x-1][6],ranks[x-1][7],rank1_idname[x-1][0],rank1_idname[x-1][1]])	 
    j=j+1
   else:
    z=j*mod	
    y=(j+1)*mod-1
    y=str(z)+"-"+str(y)		 
    rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"",y,octants_ct1[x-2],octants_ct2[x-2],octants_ct3[x-2],octants_ct4[x-2],octants_ct5[x-2],octants_ct6[x-2],octants_ct7[x-2],octants_ct8[x-2],ranks[x-1][0],ranks[x-1][1],ranks[x-1][2],ranks[x-1][3],ranks[x-1][4],ranks[x-1][5],ranks[x-1][6],ranks[x-1][7],rank1_idname[x-1][0],rank1_idname[x-1][1]])	 
    j=j+1
  elif x==k+5:
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","","Octant ID","Octant Name","Count of Rank 1 Mod values","","","",""])
  elif x==k+6:
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","","1","Internal Outward Interaction",sn_1,"","","",""])
  elif x==k+7:
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","","-1","External Outward Interaction",sn_2,"","","",""])
  elif x==k+8:
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","","2","External Ejection",sn_3,"","","",""])
  elif x==k+9:
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","","-2","Internal Ejection",sn_4,"","","",""])
  elif x==k+10:
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","","3","External inward Interaction",sn_5,"","","",""])
  elif x==k+11:
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","","-3","Internal inward Interaction",sn_6,"","","",""])
  elif x==k+12:
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","","4","Internal Sweep",sn_7,"","","",""]) 
  elif x==k+13:
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","","-4","External Sweep",sn_8,"","","",""])     
  else:
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","","","","","","","",""])

 for i in rows:
  spreadsheet.append(i)
 book.save("octant_output_ranking_excel.xlsx")
      
 octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}




mod=5000 
octant_range_names(mod)



#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))