from datetime import datetime
start_time = datetime.now()
def octant_longest_subsequence_count_with_range():
 from openpyxl import load_workbook
 import os
 import numpy as np
 os.system('cls')

 d1 = []    # declaring new list
 b1 = []    # declaring new list
 a1 = []    # declaring new list
 c1 = []    # declaring new list
 b2 = []    # declaring list for difference 
 a2 = []    # declaring list for difference
 c2 = []    # declaring list for difference
 octants = []   # declaring list for octant storing

 work_book=load_workbook("input_octant_longest_subsequence_with_range.xlsx")
 spreadsheet=work_book["Sheet1"] 
 num_1=spreadsheet.max_row
 for i in range(2,num_1+1):
  d1.append(spreadsheet.cell(row=i,column=1).value)
  a1.append(spreadsheet.cell(row=i,column=2).value)
  b1.append(spreadsheet.cell(row=i,column=3).value)
  c1.append(spreadsheet.cell(row=i,column=4).value)

 b_mean_value=np.mean(b1, dtype=np.float64)     # mean of b1
 a_mean_value=np.mean(a1, dtype=np.float64)     # mean of a1
 c_mean_value=np.mean(c1, dtype=np.float64)     # mean of c1

    #appending the differences..............
 for i in range(2,num_1+1):
  a2.append(float(spreadsheet.cell(row=i,column=2).value)-a_mean_value) 
  b2.append(float(spreadsheet.cell(row=i,column=3).value)-b_mean_value)
  c2.append(float(spreadsheet.cell(row=i,column=4).value)-c_mean_value) 

    #Finding the octant with the help of for loop
 for i in range(0,num_1-1):
  if ((a2[i]>=0) and (b2[i]>=0) and (c2[i]>=0 )):
   octants.append(1)
  elif((a2[i]>=0) and (b2[i]>=0) and (c2[i]<0 )):
   octants.append(-1)
  elif((a2[i]<0) and (b2[i]>=0) and (c2[i]>=0 )):
   octants.append(2)
  elif((a2[i]<0) and (b2[i]>=0) and (c2[i]<0 )):
   octants.append(-2)
  elif((a2[i]>=0) and (b2[i]<0) and (c2[i]>=0 )):
   octants.append(4)
  elif((a2[i]>=0) and (b2[i]<0) and (c2[i]<0 )):
   octants.append(-4)    
  elif((a2[i]<0) and (b2[i]<0) and (c2[i]>=0 )):
   octants.append(3)
  else:
   octants.append(-3)

 lss1=0         # storing the longest subsequence value of octant 1
 lss2=0         # storing the longest subsequence value of octant -1
 lss3=0         # storing the longest subsequence value of octant 2
 lss4=0         # storing the longest subsequence value of octant -2
 lss5=0         # storing the longest subsequence value of octant 3
 lss6=0         # storing the longest subsequence value of octant -3
 lss7=0         # storing the longest subsequence value of octant 4
 lss8=0         # storing the longest subsequence value of octant -4

 lssl1=0        # for storing length of the longest subsequence length for octant 1
 lssl2=0        # for storing length of the longest subsequence length for octant -1
 lssl3=0        # for storing length of the longest subsequence length for octant 2
 lssl4=0        # for storing length of the longest subsequence length for octant -2
 lssl5=0        # for storing length of the longest subsequence length for octant 3
 lssl6=0        # for storing length of the longest subsequence length for octant -3
 lssl7=0        # for storing length of the longest subsequence length for octanlss8 4
 lssl8=0        # for storing length of the longest subsequence length for octant -4

 ct1=0          # counting of octant 1
 ct2=0          # counting of octant -1
 ct3=0          # counting of octant 2
 ct4=0          # counting of octant -2
 ct5=0          # counting of octant 3
 ct6=0          # counting of octant -3
 ct7=0          # counting of octant 4
 ct8=0          # counting of octant -4

  # using for loop to get the count and length of longest subsequence for octant 1
 for i in range(0,num_1-1):
  if octants[i]==1 :
   lss1=lss1+1
   if i==num_1-2:
    if lss1>lssl1:
     lssl1=lss1
     lss1=0
     ct1=1
    elif lssl1>lss1:
     lss1=0
    else:
     lss1=0
     ct1=ct1+1 
  else:
    if lss1>lssl1:
     lssl1=lss1
     lss1=0
     ct1=1
    elif lssl1>lss1:
     lss1=0
    else:
     lss1=0
     ct1=ct1+1

 count1=0 

 list1=[]
 for i in range(0,num_1-1):  
  if octants[i]==1:
   if count1==0: 
    list_1=d1[i] 
   count1=count1+1
  else:
   count1=0 
  if count1==lssl1:
   list1.append(list_1); 
   count1=0

#  print(list1)

      # using for loop to get the count and length of longest subsequence for octant -1
 
 for i in range(0,num_1-1):
  if octants[i]==-1 :
   lss2=lss2+1
   if i==num_1-2:
    if lss2>lssl2:
     lssl2=lss2
     lss2=0
     ct2=1
    elif lssl2>lss2:
     lss2=0
    else:
     lss2=0
     ct2=ct2+1 
  else:
    if lss2>lssl2:
     lssl2=lss2
     lss2=0
     ct2=1
    elif lssl2>lss2:
     lss2=0
    else:
     lss2=0
     ct2=ct2+1

 count1=0

 list2=[]
 for i in range(0,num_1-1):  
  if octants[i]==-1:
   if count1==0: 
    list_2=d1[i] 
   count1=count1+1
  else:
   count1=0 
  if count1==lssl2:
   list2.append(list_2); 
   count1=0

#  print(list2) 
       

      # using for loop to get the count and length of longest subsequence for octant 2

 for i in range(0,num_1-1):
  if octants[i]==2 :
   lss3=lss3+1
   if i==num_1-2:
    if lss3>lssl3:
     lssl3=lss3
     lss3=0
     ct3=1
    elif lssl3>lss3:
     lss3=0
    else:
     lss3=0
     ct3=ct3+1 
  else:
    if lss3>lssl3:
     lssl3=lss3
     lss3=0
     ct3=1
    elif lssl3>lss3:
     lss3=0
    else:
     lss3=0
     ct3=ct3+1    
 
 count1=0

 list3=[]
 for i in range(0,num_1-1):  
  if octants[i]==2:
   if count1==0: 
    list_3=d1[i] 
   count1=count1+1
  else:
   count1=0 
  if count1==lssl3:
   list3.append(list_3); 
   count1=0

#  print(list3)  
 
      # using for loop to get the count and length of longest subsequence for octant -2
 
 for i in range(0,num_1-1):
  if octants[i]==-2 :
   lss4=lss4+1
   if i==num_1-2:
    if lss4>lssl4:
     lssl4=lss4
     lss4=0
     ct4=1
    elif lssl4>lss4:
     lss4=0
    else:
     lss4=0
     ct4=ct4+1 
  else:
    if lss4>lssl4:
     lssl4=lss4
     lss4=0
     ct4=1 
    elif lssl4>lss4:
     lss4=0
    else:
     lss4=0
     ct4=ct4+1    

 count1=0

 list4=[]
 for i in range(0,num_1-1):  
  if octants[i]==-2:
   if count1==0: 
    list_4=d1[i] 
   count1=count1+1
  else:
   count1=0 
  if count1==lssl4:
   list4.append(list_4); 
   count1=0

#  print(list4)  
     
      # using for loop to get the count and length of longest subsequence for octant 3

 for i in range(0,num_1-1):
  if octants[i]==3 :
   lss5=lss5+1
   if i==num_1-2:
    if lss5>lssl5:
     lssl5=lss5
     lss5=0
     ct5=1
    elif lssl5>lss5:
     lss5=0
    else:
     lss5=0
     ct5=ct5+1 
  else:
    if lss5>lssl5:
     lssl5=lss5
     lss5=0
     ct5=1
    elif lssl5>lss5:
     lss5=0
    else:
     lss5=0
     ct5=ct5+1    
 
 count1=0

 list5=[]
 for i in range(0,num_1-1):  
  if octants[i]==3:
   if count1==0: 
    list_5=d1[i] 
   count1=count1+1
  else:
   count1=0 
  if count1==lssl5:
   list5.append(list_5); 
   count1=0

#  print(list5)  
 
        # using for loop to get the count and length of longest subsequence for octant -3

 for i in range(0,num_1-1):
  if octants[i]==-3 :
   lss6=lss6+1
   if i==num_1-2:
    if lss6>lssl1:
     lssl6=lss6
     lss6=0
     ct6=1
    elif lssl6>lss6:
     lss6=0
    else:
     lss6=0
     ct6=ct6+1 
  else:
    if lss6>lssl6:
     lssl6=lss6
     lss6=0
     ct6=1
    elif lssl6>lss6:
     lss6=0
    else:
     lss6=0
     ct6=ct6+1    
 
 count1=0

 list6=[]
 for i in range(0,num_1-1):  
  if octants[i]==-3:
   if count1==0: 
    list_6=d1[i] 
   count1=count1+1
  else:
   count1=0 
  if count1==lssl6:
   list6.append(list_6); 
   count1=0

#  print(list6)  
 

        # using for loop to get the count and length of longest subsequence for octant 4
 
 for i in range(0,num_1-1):
  if octants[i]==4 :
   lss7=lss7+1
   if i==num_1-2:
    if lss7>lssl7:
     lssl7=lss7
     lss7=0
     ct7=1
    elif lssl7>lss7:
     lss7=0
    else:
     lss7=0
     ct7=ct7+1 
  else:
    if lss7>lssl7:
     lssl7=lss7
     lss7=0
     ct7=1
    elif lssl7>lss7:
     lss7=0
    else:
     lss7=0
     ct7=ct7+1

 count1=0

 list7=[]
 for i in range(0,num_1-1):  
  if octants[i]==4:
   if count1==0: 
    list_7=d1[i] 
   count1=count1+1
  else:
   count1=0 
  if count1==lssl7:
   list7.append(list_7); 
   count1=0

#  print(list7)  
 
     # using for loop to get the count and length of longest subsequence for octant -4

 for i in range(0,num_1-1):
  if octants[i]==-4 :
   lss8=lss8+1
   if i==num_1-2:
    if lss8>lssl1:
     lssl8=lss8
     lss8=0
     ct8=1
    elif lssl8>lss8:
     lss8=0
    else:
     lss8=0
     ct8=ct8+1 
  else:
    if lss8>lssl8:
     lssl8=lss8
     lss8=0
     ct8=1
    elif lssl8>lss8:
     lss8=0
    else:
     lss8=0
     ct8=ct8+1    
       
 count1=0

 list8=[]
 for i in range(0,num_1-1):  
  if octants[i]==-4:
   if count1==0: 
    list_8=d1[i] 
   count1=count1+1
  else:
   count1=0 
  if count1==lssl8:
   list8.append(list_8); 
   count1=0

#  print(list8)

 m1=[]
 m1.append("Octant")
 m1.append("1")
 m1.append("Time")
 for i in range(0,ct1):
  m1.append("")
 m1.append("-1")
 m1.append("Time")
 for i in range(0,ct2):
  m1.append("")
 m1.append("2")
 m1.append("Time")
 for i in range(0,ct3):
  m1.append("")
 m1.append("-2")
 m1.append("Time")
 for i in range(0,ct4):
  m1.append("")
 m1.append("3")
 m1.append("Time")
 for i in range(0,ct5):
  m1.append("")
 m1.append("-3")
 m1.append("Time")
 for i in range(0,ct6):
  m1.append("")
 m1.append("4")
 m1.append("Time")
 for i in range(0,ct7):
  m1.append("")
 m1.append("-4")
 m1.append("Time")
 for i in range(0,ct8):
  m1.append("")
 v=len(m1)
 n1=[]
 n1.append("Longest Subsequence Length")
 n1.append(lssl1)
 n1.append("From")
 for i in range(0,ct1):
  n1.append(list1[i])
 n1.append(lssl2)
 n1.append("From")
 for i in range(0,ct2):
  n1.append(list2[i])
 n1.append(lssl3)
 n1.append("From")
 for i in range(0,ct3):
  n1.append(list3[i])
 n1.append(lssl4)
 n1.append("From")
 for i in range(0,ct4):
  n1.append(list4[i])
 n1.append(lssl5)
 n1.append("From")
 for i in range(0,ct5):
  n1.append(list5[i])
 n1.append(lssl6)
 n1.append("From")
 for i in range(0,ct6):
  n1.append(list6[i])
 n1.append(lssl7)
 n1.append("From")
 for i in range(0,ct7):
  n1.append(list7[i])
 n1.append(lssl8)
 n1.append("From")
 for i in range(0,ct8):
  n1.append(list8[i])

 p1=[]
 p1.append("Count")
 p1.append(ct1)
 p1.append("To")
 for i in range(0,ct1):
  p1.append(list1[i]+ 0.01*(lssl1-1))
 p1.append(ct2)
 p1.append("To")
 for i in range(0,ct2):
  p1.append(list2[i]+ 0.01*(lssl2-1))
 p1.append(ct3)
 p1.append("To")
 for i in range(0,ct3):
  p1.append(list3[i]+ 0.01*(lssl3-1))
 p1.append(ct4)
 p1.append("To")
 for i in range(0,ct4):
  p1.append(list4[i]+ 0.01*(lssl4-1))
 p1.append(ct5)
 p1.append("To")
 for i in range(0,ct5):
  p1.append(list5[i]+ 0.01*(lssl5-1))
 p1.append(ct6)
 p1.append("To")
 for i in range(0,ct6):
  p1.append(list6[i]+ 0.01*(lssl6-1))
 p1.append(ct7)
 p1.append("To")
 for i in range(0,ct7):
  p1.append(list7[i]+ 0.01*(lssl7-1))
 p1.append(ct8)
 p1.append("To")
 for i in range(0,ct8):
  p1.append(list8[i]+ 0.01*(lssl8-1))

 from openpyxl import Workbook 
 book=Workbook()
 spreadsheet= book.active    
 rows=[] 
        #Writing data in output xlsx file
 rows.append(["Time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant"])
 for x in range(num_1-2):
  if(x==0):
   rows.append([d1[x],a1[x],b1[x],c1[x],a_mean_value,b_mean_value,c_mean_value,a2[x],b2[x],c2[x],octants[x],"","Octant","Longest Susequence Length","Count","","Octant","Longest Susequence Length","Count"])
    #Writing data in output xlsx file for octant 1
  elif x==1:
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","1",lssl1,ct1,"",m1[x],n1[x],p1[x]])
    #Writing data in output xlsx file for octant -1
  elif x==2:   
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","-1",lssl2,ct2,"",m1[x],n1[x],p1[x]])
    #Writing data in output xlsx file for octant 2
  elif x==3: 
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","2",lssl3,ct3,"",m1[x],n1[x],p1[x]])
    #Writing data in output xlsx file for octant -2 
  elif x==4: 
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","-2",lssl4,ct4,"",m1[x],n1[x],p1[x]]) 
    #Writing data in output xlsx file for octant 3
  elif x==5:  
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","3",lssl5,ct5,"",m1[x],n1[x],p1[x]])
    #Writing data in output xlsx file for octant -3
  elif x==6:  
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","-3",lssl6,ct6,"",m1[x],n1[x],p1[x]])
    #Writing data in output xlsx file for octant 4 
  elif x==7: 
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","4",lssl7,ct7,"",m1[x],n1[x],p1[x]])
    #Writing data in output xlsx file for octant -4
  elif x==8: 
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","-4",lssl8,ct8,"",m1[x],n1[x],p1[x]])
  elif x<v:
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x],"","","","","",m1[x],n1[x],p1[x]])
    #writing remaining data in xlsx file
  else: 
   rows.append([d1[x],a1[x],b1[x],c1[x],"","","",a2[x],b2[x],c2[x],octants[x]])

 for i in rows:
  spreadsheet.append(i)
 book.save("output_octant_longest_subsequence_with_range.xlsx")
 
octant_longest_subsequence_count_with_range()

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
