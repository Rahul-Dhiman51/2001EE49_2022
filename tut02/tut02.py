def octant_transition_count(mod=5000):
 from openpyxl import load_workbook
 import os
 import numpy as np
 
 os.system('cls')

 d1 = []    # declaring new list
 b1 = []    # declaring new list
 a1 = []    # declaring new list
 c1 = []    # declaring new list
 b2 = []    # declaring list for difference 
 a2 = []    # declaring list for difference
 c2 = []    # declaring list for difference
 octants = []   # declaring list for octant storing

 work_book=load_workbook("input_octant_transition_identify.xlsx")
 spreadsheet=work_book["Sheet1"]
 num_1=spreadsheet.max_row

 for i in range(2,num_1+1):
  d1.append(spreadsheet.cell(row=i,column=1).value)
  a1.append(spreadsheet.cell(row=i,column=2).value)
  b1.append(spreadsheet.cell(row=i,column=3).value)
  c1.append(spreadsheet.cell(row=i,column=4).value)

 b_mean_value=np.mean(b1, dtype=np.float64) # for mean of b1
 a_mean_value=np.mean(a1, dtype=np.float64) # for mean of a1
 c_mean_value=np.mean(c1, dtype=np.float64) # for mean of c1

 for i in range(2,num_1+1):
  a2.append(float(spreadsheet.cell(row=i,column=2).value)-a_mean_value)
  b2.append(float(spreadsheet.cell(row=i,column=3).value)-b_mean_value)
  c2.append(float(spreadsheet.cell(row=i,column=4).value)-c_mean_value)  

  # Finding octant with the help of for loop
 for i in range(0,num_1-1):
  if ((a2[i]>=0) and (b2[i]>=0) and (c2[i]>=0 )):
   octants.append(1)
  elif((a2[i]>=0) and (b2[i]>=0) and (c2[i]<0 )):
   octants.append(-1)
  elif((a2[i]<0) and (b2[i]>=0) and (c2[i]>=0 )):
   octants.append(2)
  elif((a2[i]<0) and (b2[i]>=0) and (c2[i]<0 )):
   octants.append(-2)
  elif((a2[i]>=0) and (b2[i]<0) and (c2[i]>=0 )):
   octants.append(4)
  elif((a2[i]>=0) and (b2[i]<0) and (c2[i]<0 )):
   octants.append(-4)    
  elif((a2[i]<0) and (b2[i]<0) and (c2[i]>=0 )):
   octants.append(3)
  else:
   octants.append(-3) 
     
  ct1=0         # counting of octant 1
  ct2=0         # counting of octant -1
  ct3=0         # counting of octant 2
  ct4=0         # counting of octant -2
  ct5=0         # counting of octant 3
  ct6=0         # counting of octant -3
  ct7=0         # counting of octant 4
  ct8=0         # counting of octant -4

  #finding the octants using loop
 for i in range(0,num_1-1):
  if (octants[i]==1):
   ct1=ct1+1
  elif(octants[i]==-1):
   ct2=ct2+1
  elif(octants[i]==2):
   ct3=ct3+1
  elif(octants[i]==-2):
   ct4=ct4+1
  elif(octants[i]==3):
   ct5=ct5+1
  elif(octants[i]==-3):
   ct6=ct6+1    
  elif(octants[i]==4):
   ct7=ct7+1
  else:
   ct8=ct8+1

 # Making lists of different octants in mod range

 octants_ct1=[]      
 octants_ct2=[] 
 octants_ct3=[]     
 octants_ct4=[]
 octants_ct5=[]
 octants_ct6=[] 
 octants_ct7=[]
 octants_ct8=[]

  #count of octants in mod range...

 octs_1=0
 octs_2=0
 octs_3=0
 octs_4=0
 octs_5=0
 octs_6=0
 octs_7=0
 octs_8=0

 k= int((num_1-2)/mod) +1 # k gives the value of interval.

 for l in range(k):
  y=mod*l
  for t in range(mod):	
   if(t+y<num_1-1):	                                    
    if (octants[t+y]==1):
     octs_1=octs_1+1
    elif(octants[t+y]==-1):
     octs_2=octs_2+1
    elif(octants[t+y]==2):
     octs_3=octs_3+1
    elif(octants[t+y]==-2):
     octs_4=octs_4+1
    elif(octants[t+y]==3):
     octs_5=octs_5+1
    elif(octants[t+y]==-3):
     octs_6=octs_6+1    
    elif(octants[t+y]==4):
     octs_7=octs_7+1
    elif(octants[t+y]==-4):
     octs_8=octs_8+1

  octants_ct1.append(octs_1)
  octants_ct2.append(octs_2) 
  octants_ct3.append(octs_3)
  octants_ct4.append(octs_4)
  octants_ct5.append(octs_5)
  octants_ct6.append(octs_6)
  octants_ct7.append(octs_7)
  octants_ct8.append(octs_8)

  octs_1=0
  octs_2=0
  octs_3=0
  octs_4=0
  octs_5=0
  octs_6=0
  octs_7=0
  octs_8=0

 final_mat=[] # stores full data that will be printed.
 mat_2d=[]  #stores 2-d matrix data of respective interval
 overall_trans=[[0]*8,[0]*8,[0]*8,[0]*8,[0]*8,[0]*8,[0]*8,[0]*8]  #matrix of overall transition
 for m in range(0,k) :
  f=m*mod

  # Following lists stores the value of each element in 2-d matrix of different intervals
  value_list1=[0]*8
  value_list2=[0]*8
  value_list3=[0]*8
  value_list4=[0]*8
  value_list5=[0]*8
  value_list6=[0]*8
  value_list7=[0]*8
  value_list8=[0]*8


  for j in range(0,mod):
   if j+f<=num_1-3:
    if octants[j+f]==1 and octants[j+f+1]==1 :
      value_list1[0]=value_list1[0]+1   # This will keep on incresing value after each transition in different intervals.
      overall_trans[0][0]=overall_trans[0][0]+1   #For all overall transition matrix
    elif octants[j+f]==1 and octants[j+f+1]==-1:
      value_list1[1]=value_list1[1]+1    # This will keep on incresing value after each transition in different intervals.
      overall_trans[0][1]=overall_trans[0][1]+1   #For all overall transition matrix
    elif octants[j+f]==1 and octants[j+f+1]==2:
      value_list1[2]=value_list1[2]+1     # This will keep on incresing value after each transition in different intervals.
      overall_trans[0][2]=overall_trans[0][2]+1   #For all overall transition matrix
    elif octants[j+f]==1 and octants[j+f+1]==-2:
      value_list1[3]=value_list1[3]+1
      overall_trans[0][3]=overall_trans[0][3]+1
    elif octants[j+f]==1 and octants[j+1+f]==3:
      value_list1[4]=value_list1[4]+1 
      overall_trans[0][4]=overall_trans[0][4]+1
    elif octants[j+f]==1 and octants[j+f+1]==-3:
      value_list1[5]=value_list1[5]+1
      overall_trans[0][5]=overall_trans[0][5]+1
    elif octants[j+f]==1 and octants[j+1+f]==4:
      value_list1[6]=value_list1[6]+1
      overall_trans[0][6]=overall_trans[0][6]+1
    elif octants[j+f]==1 and octants[j+f+1]==-4:
      value_list1[7]=value_list1[7]+1
      overall_trans[0][7]=overall_trans[0][7]+1
  
    elif octants[j+f]==-1 and octants[j+f+1]==1 :
      value_list2[0]=value_list2[0]+1
      overall_trans[1][0]=overall_trans[1][0]+1
    elif octants[j+f]==-1 and octants[j+f+1]==-1:
      value_list2[1]=value_list2[1]+1    
      overall_trans[1][1]=overall_trans[1][1]+1
    elif octants[j+f]==-1 and octants[j+f+1]==2:
      value_list2[2]=value_list2[2]+1
      overall_trans[1][2]=overall_trans[1][2]+1
    elif octants[j+f]==-1 and octants[j+f+1]==-2:
      value_list2[3]=value_list2[3]+1
      overall_trans[1][3]=overall_trans[1][3]+1
    elif octants[j+f]==-1 and octants[j+f+1]==3:
      value_list2[4]=value_list2[4]+1
      overall_trans[1][4]=overall_trans[1][4]+1 
    elif octants[j+f]==-1 and octants[j+1+f]==-3:
      value_list2[5]=value_list2[5]+1
      overall_trans[1][5]=overall_trans[1][5]+1
    elif octants[j+f]==-1 and octants[j+f+1]==4:
      value_list2[6]=value_list2[6]+1
      overall_trans[1][6]=overall_trans[1][6]+1
    elif octants[j+f]==-1 and octants[j+1+f]==-4:
      value_list2[7]=value_list2[7]+1
      overall_trans[1][7]=overall_trans[1][7]+1
    
    elif octants[j+f]==2 and octants[j+f+1]==1 :
      value_list3[0]=value_list3[0]+1
      overall_trans[2][0]=overall_trans[2][0]+1
    elif octants[j+f]==2 and octants[j+f+1]==-1:
      value_list3[1]=value_list3[1]+1
      overall_trans[2][1]=overall_trans[2][1]+1    
    elif octants[j+f]==2 and octants[j+f+1]==2:
      value_list3[2]=value_list3[2]+1
      overall_trans[2][2]=overall_trans[2][2]+1
    elif octants[j+f]==2 and octants[j+f+1]==-2:
      value_list3[3]=value_list3[3]+1
      overall_trans[2][3]=overall_trans[2][3]+1
    elif octants[j+f]==2 and octants[j+f+1]==3:
      value_list3[4]=value_list3[4]+1
      overall_trans[2][4]=overall_trans[2][4]+1 
    elif octants[j+f]==2 and octants[j+1+f]==-3:
      value_list3[5]=value_list3[5]+1
      overall_trans[2][5]=overall_trans[2][5]+1
    elif octants[j+f]==2 and octants[j+f+1]==4:
      value_list3[6]=value_list3[6]+1
      overall_trans[2][6]=overall_trans[2][6]+1
    elif octants[j+f]==2 and octants[j+1+f]==-4:
      value_list3[7]=value_list3[7]+1
      overall_trans[2][7]=overall_trans[2][7]+1

    elif octants[j+f]==-2 and octants[j+f+1]==1 :
      value_list4[0]=value_list4[0]+1
      overall_trans[3][0]=overall_trans[3][0]+1
    elif octants[j+f]==-2 and octants[j+f+1]==-1:
      value_list4[1]=value_list4[1]+1
      overall_trans[3][1]=overall_trans[3][1]+1    
    elif octants[j+f]==-2 and octants[j+f+1]==2:
      value_list4[2]=value_list4[2]+1
      overall_trans[3][2]=overall_trans[3][2]+1
    elif octants[j+f]==-2 and octants[j+f+1]==-2:
      value_list4[3]=value_list4[3]+1
      overall_trans[3][3]=overall_trans[3][3]+1
    elif octants[j+f]==-2 and octants[j+f+1]==3:
      value_list4[4]=value_list4[4]+1 
      overall_trans[3][4]=overall_trans[3][4]+1
    elif octants[j+f]==-2 and octants[j+1+f]==-3:
      value_list4[5]=value_list4[5]+1
      overall_trans[3][5]=overall_trans[3][5]+1
    elif octants[j+f]==-2 and octants[j+f+1]==4:
      value_list4[6]=value_list4[6]+1
      overall_trans[3][6]=overall_trans[3][6]+1
    elif octants[j+f]==-2 and octants[j+1+f]==-4:
      value_list4[7]=value_list4[7]+1
      overall_trans[3][7]=overall_trans[3][7]+1  
  
    elif octants[j+f]==3 and octants[j+f+1]==1 :
      value_list5[0]=value_list5[0]+1
      overall_trans[4][0]=overall_trans[4][0]+1
    elif octants[j+f]==3 and octants[j+f+1]==-1:
      value_list5[1]=value_list5[1]+1
      overall_trans[4][1]=overall_trans[4][1]+1    
    elif octants[j+f]==3 and octants[j+f+1]==2:
      value_list5[2]=value_list5[2]+1
      overall_trans[4][2]=overall_trans[4][2]+1
    elif octants[j+f]==3 and octants[j+f+1]==-2:
      value_list5[3]=value_list5[3]+1
      overall_trans[4][3]=overall_trans[4][3]+1
    elif octants[j+f]==3 and octants[j+f+1]==3:
      value_list5[4]=value_list5[4]+1 
      overall_trans[4][4]=overall_trans[4][4]+1
    elif octants[j+f]==3 and octants[j+1+f]==-3:
      value_list5[5]=value_list5[5]+1
      overall_trans[4][5]=overall_trans[4][5]+1
    elif octants[j+f]==3 and octants[j+f+1]==4:
      value_list5[6]=value_list5[6]+1
      overall_trans[4][6]=overall_trans[4][6]+1
    elif octants[j+f]==3 and octants[j+1+f]==-4:
      value_list5[7]=value_list5[7]+1
      overall_trans[4][7]=overall_trans[4][7]+1 
    
    elif octants[j+f]==-3 and octants[j+f+1]==1 :
      value_list6[0]=value_list6[0]+1
      overall_trans[5][0]=overall_trans[5][0]+1
    elif octants[j+f]==-3 and octants[j+f+1]==-1:
      value_list6[1]=value_list6[1]+1    
      overall_trans[5][1]=overall_trans[5][1]+1
    elif octants[j+f]==-3 and octants[j+f+1]==2:
      value_list6[2]=value_list6[2]+1
      overall_trans[5][2]=overall_trans[5][2]+1
    elif octants[j+f]==-3 and octants[j+f+1]==-2:
      value_list6[3]=value_list6[3]+1
      overall_trans[5][3]=overall_trans[5][3]+1
    elif octants[j+f]==-3 and octants[j+f+1]==3:
      value_list6[4]=value_list6[4]+1
      overall_trans[5][4]=overall_trans[5][4]+1 
    elif octants[j+f]==-3 and octants[j+1+f]==-3:
      value_list6[5]=value_list6[5]+1
      overall_trans[5][5]=overall_trans[5][5]+1
    elif octants[j+f]==-3 and octants[j+f+1]==4:
      value_list6[6]=value_list6[6]+1
      overall_trans[5][6]=overall_trans[5][6]+1
    elif octants[j+f]==-3 and octants[j+1+f]==-4:
      value_list6[7]=value_list6[7]+1
      overall_trans[5][7]=overall_trans[5][7]+1

    elif octants[j+f]==4 and octants[j+f+1]==1 :
      value_list7[0]=value_list7[0]+1
      overall_trans[6][0]=overall_trans[6][0]+1
    elif octants[j+f]==4 and octants[j+f+1]==-1:
      value_list7[1]=value_list7[1]+1    
      overall_trans[6][1]=overall_trans[6][1]+1
    elif octants[j+f]==4 and octants[j+f+1]==2:
      value_list7[2]=value_list7[2]+1
      overall_trans[6][2]=overall_trans[6][2]+1
    elif octants[j+f]==4 and octants[j+f+1]==-2:
      value_list7[3]=value_list7[3]+1
      overall_trans[6][3]=overall_trans[6][3]+1
    elif octants[j+f]==4 and octants[j+f+1]==3:
      value_list7[4]=value_list7[4]+1
      overall_trans[6][4]=overall_trans[6][4]+1
    elif octants[j+f]==4 and octants[j+1+f]==-3:
      value_list7[5]=value_list7[5]+1
      overall_trans[6][5]=overall_trans[6][5]+1
    elif octants[j+f]==4 and octants[j+f+1]==4:
      value_list7[6]=value_list7[6]+1
      overall_trans[6][6]=overall_trans[6][6]+1
    elif octants[j+f]==4 and octants[j+1+f]==-4:
      value_list7[7]=value_list7[7]+1 
      overall_trans[6][7]=overall_trans[6][7]+1
    elif octants[j+f]==-4 and octants[j+f+1]==1 :
      value_list8[0]=value_list8[0]+1
      overall_trans[7][0]=overall_trans[7][0]+1
    elif octants[j+f]==-4 and octants[j+f+1]==-1:
      value_list8[1]=value_list8[1]+1 
      overall_trans[7][1]=overall_trans[7][1]+1  
    elif octants[j+f]==-4 and octants[j+f+1]==2:
      value_list8[2]=value_list8[2]+1
      overall_trans[7][2]=overall_trans[7][2]+1
    elif octants[j+f]==-4 and octants[j+f+1]==-2:
      value_list8[3]=value_list8[3]+1
      overall_trans[7][3]=overall_trans[7][3]+1
    elif octants[j+f]==-4 and octants[j+f+1]==3:
      value_list8[4]=value_list8[4]+1
      overall_trans[7][4]=overall_trans[7][4]+1 
    elif octants[j+f]==-4 and octants[j+1+f]==-3:
      value_list8[5]=value_list8[5]+1
      overall_trans[7][5]=overall_trans[7][5]+1
    elif octants[j+f]==-4 and octants[j+f+1]==4:
      value_list8[6]=value_list8[6]+1
      overall_trans[7][6]=overall_trans[7][6]+1
    elif octants[j+f]==-4 and octants[j+1+f]==-4:
      value_list8[7]=value_list8[7]+1  
      overall_trans[7][7]=overall_trans[7][7]+1

  if m==0 :

    final_mat.append(overall_trans)
    # Appending eight 1-d matrix in mat_2d 2-d matrix
  mat_2d.append(value_list1)  
  mat_2d.append(value_list2) 
  mat_2d.append(value_list3)
  mat_2d.append(value_list4)
  mat_2d.append(value_list5)
  mat_2d.append(value_list6)
  mat_2d.append(value_list7)
  mat_2d.append(value_list8)
  grid=mat_2d.copy()

  final_mat.append(grid)      # appending mat_2d in final_mat

  mat_2d.clear()
 h = ["+1","-1","+2","-2","+3","-3","+4","-4"]

  # writing code for printing the output......
 from openpyxl import Workbook 
 book=Workbook()
 spreadsheet= book.active    
 rows=[] 
 rows.append(["Time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant","","OctantID","+1","-1","+2","-2","+3","-3","+4","-4"])
 j=0
 a=0
 b=0
 for q in range(num_1-2):
  if(q==0):
   rows.append([d1[q],a1[q],b1[q],c1[q],a_mean_value,b_mean_value,c_mean_value,a2[q],b2[q],c2[q],octants[q],"","Overall count",ct1,ct2,ct3,ct4,ct5,ct6,ct7,ct8])
  elif(q==1):
   st="mod "+str(mod)		
   rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"User input",st,"","","","","","","",""])
  elif(q>=2 and q<2+k):
   if(q==1+k):
    x=j*mod 	
    z=(j+1)*mod
    st=str(x)+"-"+str(num_1-2)		 
    rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"",st,octants_ct1[q-2],octants_ct2[q-2],octants_ct3[q-2],octants_ct4[q-2],octants_ct5[q-2],octants_ct6[q-2],octants_ct7[q-2],octants_ct8[q-2]])	 
    rows.append([d1[q],a1[q],b1[q],c1[q],a_mean_value,b_mean_value,c_mean_value,a2[q],b2[q],c2[q],octants[q],"","Verified",ct1,ct2,ct3,ct4,ct5,ct6,ct7,ct8])
    j=j+1
   else:
    x=j*mod	
    z=(j+1)*mod-1
    st=str(x)+"-"+str(z)		 
    rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"",st,octants_ct1[q-2],octants_ct2[q-2],octants_ct3[q-2],octants_ct4[q-2],octants_ct5[q-2],octants_ct6[q-2],octants_ct7[q-2],octants_ct8[q-2]])	 
    j=j+1
  elif (q-(2+k))%9==0 and q<9*(k+1)+2+k:
    rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","","","","","","","","",""])
    rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","","","","","","","","",""])
    if q ==2+k:
     rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","Overall transition Count","","","","","","","",""])
     rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","","To","","","","","","",""])
    else:
     rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","Mod transition Count","","","","","","","",""])
     rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"",str(((q-(2+k))//9-1)*mod)+"-"+str(np.minimum(((q-(2+k))//9)*mod,num_1-2)),"To","","","","","","",""])
    rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","Count","+1","-1","+2","-2","+3","-3","+4","-4"])
  elif q<9*(k+1)+(2+k):
   if (q-(2+k))%9==1:
    rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"From",h[((q-(2+k))%9)-1],final_mat[b][a][0],final_mat[b][a][1],final_mat[b][a][2],final_mat[b][a][3],final_mat[b][a][4],final_mat[b][a][5],final_mat[b][a][6],final_mat[b][a][7]])	
   else:
    rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"",h[((q-(2+k))%9)-1],final_mat[b][a][0],final_mat[b][a][1],final_mat[b][a][2],final_mat[b][a][3],final_mat[b][a][4],final_mat[b][a][5],final_mat[b][a][6],final_mat[b][a][7]])
   a=a+1
   if a==8:
    a=0
    b=b+1
  else:
   rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","","","","","","","",""])

 for i in rows:
  spreadsheet.append(i)
 book.save("output_octant_transition_identify.xlsx")

mod=5000    # Input given here
octant_transition_count(mod)