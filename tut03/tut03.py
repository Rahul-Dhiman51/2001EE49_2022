
def octant_longest_subsequence_count():
 from openpyxl import load_workbook
 import os
 import numpy as np

 os.system('cls')

 d1 = []    # declaring new list
 b1 = []    # declaring new list
 a1 = []    # declaring new list
 c1 = []    # declaring new list
 b2 = []    # declaring list for difference 
 a2 = []    # declaring list for difference
 c2 = []    # declaring list for difference
 octants = []   # declaring list for octant storing

 work_book=load_workbook("input_octant_longest_subsequence.xlsx")
 spreadsheet=work_book["Sheet1"]
 num_1=spreadsheet.max_row
 for i in range(2,num_1+1):
  d1.append(spreadsheet.cell(row=i,column=1).value)
  a1.append(spreadsheet.cell(row=i,column=2).value)
  b1.append(spreadsheet.cell(row=i,column=3).value)
  c1.append(spreadsheet.cell(row=i,column=4).value)

 b_mean_value=np.mean(b1, dtype=np.float64)     # mean of b1
 a_mean_value=np.mean(a1, dtype=np.float64)     # mean of a1
 c_mean_value=np.mean(c1, dtype=np.float64)     # mean of c1

    #appending the differences..............
 for i in range(2,num_1+1):
  a2.append(float(spreadsheet.cell(row=i,column=2).value)-a_mean_value) 
  b2.append(float(spreadsheet.cell(row=i,column=3).value)-b_mean_value)
  c2.append(float(spreadsheet.cell(row=i,column=4).value)-c_mean_value) 

    #Finding the octant with the help of for loop
 for i in range(0,num_1-1):
  if ((a2[i]>=0) and (b2[i]>=0) and (c2[i]>=0 )):
   octants.append(1)
  elif((a2[i]>=0) and (b2[i]>=0) and (c2[i]<0 )):
   octants.append(-1)
  elif((a2[i]<0) and (b2[i]>=0) and (c2[i]>=0 )):
   octants.append(2)
  elif((a2[i]<0) and (b2[i]>=0) and (c2[i]<0 )):
   octants.append(-2)
  elif((a2[i]>=0) and (b2[i]<0) and (c2[i]>=0 )):
   octants.append(4)
  elif((a2[i]>=0) and (b2[i]<0) and (c2[i]<0 )):
   octants.append(-4)    
  elif((a2[i]<0) and (b2[i]<0) and (c2[i]>=0 )):
   octants.append(3)
  else:
   octants.append(-3)

 lss1=0         # storing the longest subsequence value of octant 1
 lss2=0         # storing the longest subsequence value of octant -1
 lss3=0         # storing the longest subsequence value of octant 2
 lss4=0         # storing the longest subsequence value of octant -2
 lss5=0         # storing the longest subsequence value of octant 3
 lss6=0         # storing the longest subsequence value of octant -3
 lss7=0         # storing the longest subsequence value of octant 4
 lss8=0         # storing the longest subsequence value of octant -4

 lssl1=0        # for storing length of the longest subsequence length for octant 1
 lssl2=0        # for storing length of the longest subsequence length for octant -1
 lssl3=0        # for storing length of the longest subsequence length for octant 2
 lssl4=0        # for storing length of the longest subsequence length for octant -2
 lssl5=0        # for storing length of the longest subsequence length for octant 3
 lssl6=0        # for storing length of the longest subsequence length for octant -3
 lssl7=0        # for storing length of the longest subsequence length for octanlss8 4
 lssl8=0        # for storing length of the longest subsequence length for octant -4

 ct1=0          # counting of octant 1
 ct2=0          # counting of octant -1
 ct3=0          # counting of octant 2
 ct4=0          # counting of octant -2
 ct5=0          # counting of octant 3
 ct6=0          # counting of octant -3
 ct7=0          # counting of octant 4
 ct8=0          # counting of octant -4

  # using for loop to get the count and length of longest subsequence for octant 1
 for i in range(0,num_1-1):
  if octants[i]==1 :
   lss1=lss1+1
   if i==num_1-2:
    if lss1>lssl1:
     lssl1=lss1
     lss1=0
     ct1=1
    elif lssl1>lss1:
     lss1=0
    else:
     lss1=0
     ct1=ct1+1 
  else:
    if lss1>lssl1:
     lssl1=lss1
     lss1=0
     ct1=1
    elif lssl1>lss1:
     lss1=0
    else:
     lss1=0
     ct1=ct1+1
 
      # using for loop to get the count and length of longest subsequence for octant -1
 
 for i in range(0,num_1-1):
  if octants[i]==-1 :
   lss2=lss2+1
   if i==num_1-2:
    if lss2>lssl2:
     lssl2=lss2
     lss2=0
     ct2=1
    elif lssl2>lss2:
     lss2=0
    else:
     lss2=0
     ct2=ct2+1 
  else:
    if lss2>lssl2:
     lssl2=lss2
     lss2=0
     ct2=1
    elif lssl2>lss2:
     lss2=0
    else:
     lss2=0
     ct2=ct2+1   

      # using for loop to get the count and length of longest subsequence for octant 2

 for i in range(0,num_1-1):
  if octants[i]==2 :
   lss3=lss3+1
   if i==num_1-2:
    if lss3>lssl3:
     lssl3=lss3
     lss3=0
     ct3=1
    elif lssl3>lss3:
     lss3=0
    else:
     lss3=0
     ct3=ct3+1 
  else:
    if lss3>lssl3:
     lssl3=lss3
     lss3=0
     ct3=1
    elif lssl3>lss3:
     lss3=0
    else:
     lss3=0
     ct3=ct3+1    

      # using for loop to get the count and length of longest subsequence for octant -2
 
 for i in range(0,num_1-1):
  if octants[i]==-2 :
   lss4=lss4+1
   if i==num_1-2:
    if lss4>lssl4:
     lssl4=lss4
     lss4=0
     ct4=1
    elif lssl4>lss4:
     lss4=0
    else:
     lss4=0
     ct4=ct4+1 
  else:
    if lss4>lssl4:
     lssl4=lss4
     lss4=0
     ct4=1 
    elif lssl4>lss4:
     lss4=0
    else:
     lss4=0
     ct4=ct4+1    
     

      # using for loop to get the count and length of longest subsequence for octant 3

 for i in range(0,num_1-1):
  if octants[i]==3 :
   lss5=lss5+1
   if i==num_1-2:
    if lss5>lssl5:
     lssl5=lss5
     lss5=0
     ct5=1
    elif lssl5>lss5:
     lss5=0
    else:
     lss5=0
     ct5=ct5+1 
  else:
    if lss5>lssl5:
     lssl5=lss5
     lss5=0
     ct5=1
    elif lssl5>lss5:
     lss5=0
    else:
     lss5=0
     ct5=ct5+1    


        # using for loop to get the count and length of longest subsequence for octant -3

 for i in range(0,num_1-1):
  if octants[i]==-3 :
   lss6=lss6+1
   if i==num_1-2:
    if lss6>lssl1:
     lssl6=lss6
     lss6=0
     ct6=1
    elif lssl6>lss6:
     lss6=0
    else:
     lss6=0
     ct6=ct6+1 
  else:
    if lss6>lssl6:
     lssl6=lss6
     lss6=0
     ct6=1
    elif lssl6>lss6:
     lss6=0
    else:
     lss6=0
     ct6=ct6+1    

        # using for loop to get the count and length of longest subsequence for octant 4
 
 for i in range(0,num_1-1):
  if octants[i]==4 :
   lss7=lss7+1
   if i==num_1-2:
    if lss7>lssl7:
     lssl7=lss7
     lss7=0
     ct7=1
    elif lssl7>lss7:
     lss7=0
    else:
     lss7=0
     ct7=ct7+1 
  else:
    if lss7>lssl7:
     lssl7=lss7
     lss7=0
     ct7=1
    elif lssl7>lss7:
     lss7=0
    else:
     lss7=0
     ct7=ct7+1

     # using for loop to get the count and length of longest subsequence for octant -4

 for i in range(0,num_1-1):
  if octants[i]==-4 :
   lss8=lss8+1
   if i==num_1-2:
    if lss8>lssl1:
     lssl8=lss8
     lss8=0
     ct8=1
    elif lssl8>lss8:
     lss8=0
    else:
     lss8=0
     ct8=ct8+1 
  else:
    if lss8>lssl8:
     lssl8=lss8
     lss8=0
     ct8=1
    elif lssl8>lss8:
     lss8=0
    else:
     lss8=0
     ct8=ct8+1    
  
#  print(lssl8,ct8)

 from openpyxl import Workbook 
 book=Workbook()
 spreadsheet= book.active    
 rows=[] 
 rows.append(["Time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant"])
 
 for q in range(num_1-2):
  if(q==0):
   rows.append([d1[q],a1[q],b1[q],c1[q],a_mean_value,b_mean_value,c_mean_value,a2[q],b2[q],c2[q],octants[q],"","Octant","Longest Susequence Length","Count"])
  elif q==1:  
   rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","1",lssl1,ct1])
  elif q==2:  
   rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","-1",lssl2,ct2])
  elif q==3:  
   rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","2",lssl3,ct3])
  elif q==4:  
   rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","-2",lssl4,ct4]) 
  elif q==5:  
   rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","3",lssl5,ct5])
  elif q==6:  
   rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","-3",lssl6,ct6])
  elif q==7:  
   rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","4",lssl7,ct7])
  elif q==8:  
   rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q],"","-4",lssl8,ct8])
  else:
   rows.append([d1[q],a1[q],b1[q],c1[q],"","","",a2[q],b2[q],c2[q],octants[q]])

 for i in rows:
  spreadsheet.append(i)
 book.save("output_octant_longest_subsequence.xlsx")


octant_longest_subsequence_count()