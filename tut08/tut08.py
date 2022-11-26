import openpyxl
import pandas as pd
import os
from datetime import datetime
start_time = datetime.now()
india_innings = open("india_inns2.txt","r+")    #opening india innings file
pakistan_innings = open("pak_inns1.txt","r+")   #opening pakistan innings file
teamss = open("teams.txt","r+")
input_teams = teamss.readlines()

pakistan_team = input_teams[0]
pakistan_players = pakistan_team[23:-1:].split(",")

india_team = input_teams[2]
india_players = india_team[20:-1:].split(",")


first_india=india_innings.readlines()   #reading india innings file
for i in first_india:
    if i=='\n':
        first_india.remove(i)

first_pakistan=pakistan_innings.readlines()     #reading pakistan innings file
for i in first_pakistan:
    if i=='\n':
        first_pakistan.remove(i)

wb = openpyxl.Workbook()
spreadsheet = wb.active

    #   batting [runs,ball,4s,6s,sr]
    #   bowling [over,medan,runs,Wickets, NB, WD, ECO]
india_fall_of_wickets=0
pakistan_fall_of_wickets=0
out_pakistan_bat={}
india_bowlers={}
india_bats={}

pakistan_bats={}
pakistan_bowlers={}
pakistan_byes=0
pakistan_bowlers_total=0
for k in first_pakistan:
    t=k.index(".")
    overs_pakistan=k[0:t+2]
    temp=k[t+2::].split(",")
    current_ball=temp[0].split("to")


	if f"{current_ball[0].strip()}" not in india_bowlers.keys() :
        india_bowlers[f"{current_ball[0].strip()}"]=[1,0,0,0,0,0,0]     #   [over0, medan1, runs2, Wickets3, NB4, WD5, ECO6]
    elif "wide" in temp[1]:
        pass
    elif "byes" in temp[1]:
        if "FOUR" in temp[2]:
            pakistan_byes+=4
            india_bowlers[f"{current_ball[0].strip()}"][0]+=1
        elif "1 run" in temp[2]:
            pakistan_byes+=1
            india_bowlers[f"{current_ball[0].strip()}"][0]+=1
        elif "2 runs" in temp[2]:
            pakistan_byes+=2
            india_bowlers[f"{current_ball[0].strip()}"][0]+=1
        elif "3 runs" in temp[2]:
            pakistan_byes+=3
            india_bowlers[f"{current_ball[0].strip()}"][0]+=1
        elif "4 runs" in temp[2]:
            pakistan_byes+=4
            india_bowlers[f"{current_ball[0].strip()}"][0]+=1
        elif "5 runs" in temp[2]:
            pakistan_byes+=5
            india_bowlers[f"{current_ball[0].strip()}"][0]+=1

    else:
        india_bowlers[f"{current_ball[0].strip()}"][0]+=1
    
    if f"{current_ball[1].strip()}" not in pakistan_bats.keys() and temp[1]!="wide":
        pakistan_bats[f"{current_ball[1].strip()}"]=[0,1,0,0,0]     #   [runs, ball, 4s, 6s, sr]
    elif "wide" in temp[1] :
        pass
    else:
        pakistan_bats[f"{current_ball[1].strip()}"][1]+=1
    

    if "out" in temp[1]:
        india_bowlers[f"{current_ball[0].strip()}"][3]+=1
        if "Bowled" in temp[1].split("!!")[0]:
            out_pakistan_bat[f"{current_ball[1].strip()}"]=("b" + current_ball[0])
        elif "Caught" in temp[1].split("!!")[0]:
            w=(temp[1].split("!!")[0]).split("by")
            out_pakistan_bat[f"{current_ball[1].strip()}"]=("c" + w[1] +" b " + current_ball[0])
        elif "Lbw" in temp[1].split("!!")[0]:
            out_pakistan_bat[f"{current_ball[1].strip()}"]=("lbw  b "+current_ball[0])

    

    if "no run" in temp[1] or "out" in temp[1] :
        india_bowlers[f"{current_ball[0].strip()}"][2]+=0
        pakistan_bats[f"{current_ball[1].strip()}"][0]+=0
    elif "1 run" in temp[1]:
        india_bowlers[f"{current_ball[0].strip()}"][2]+=1
        pakistan_bats[f"{current_ball[1].strip()}"][0]+=1
    elif "2 runs" in temp[1]:
        india_bowlers[f"{current_ball[0].strip()}"][2]+=2
        pakistan_bats[f"{current_ball[1].strip()}"][0]+=2
    elif "3 runs" in temp[1]:
        india_bowlers[f"{current_ball[0].strip()}"][2]+=3
        pakistan_bats[f"{current_ball[1].strip()}"][0]+=3
    elif "4 runs" in temp[1]:
        india_bowlers[f"{current_ball[0].strip()}"][2]+=4
        pakistan_bats[f"{current_ball[1].strip()}"][0]+=4
    elif "FOUR" in temp[1]:
        india_bowlers[f"{current_ball[0].strip()}"][2]+=4
        pakistan_bats[f"{current_ball[1].strip()}"][0]+=4
        pakistan_bats[f"{current_ball[1].strip()}"][2]+=1
    elif "SIX" in temp[1]:
        india_bowlers[f"{current_ball[0].strip()}"][2]+=6
        pakistan_bats[f"{current_ball[1].strip()}"][0]+=6
        pakistan_bats[f"{current_ball[1].strip()}"][3]+=1
    elif "wide" in temp[1]:
        if "wides" in temp[1]:

            india_bowlers[f"{current_ball[0].strip()}"][2]+=int(temp[1][1])
            india_bowlers[f"{current_ball[0].strip()}"][5]+=int(temp[1][1])
        else:
            india_bowlers[f"{current_ball[0].strip()}"][2]+=1
            india_bowlers[f"{current_ball[0].strip()}"][5]+=1

for value in pakistan_bats.values():
    value[-1]=round((value[0]/value[1])*100 , 2)


############# india innings ##############


india_bowlers_total=0
india_byes=0

out_india_bat={}
for k in first_india:
    t=k.index(".")
    overs_india=k[0:t+2]

    temp=k[t+2::].split(",")

    current_ball=temp[0].split("to") #0 2
    if f"{current_ball[0].strip()}" not in pakistan_bowlers.keys() :
        pakistan_bowlers[f"{current_ball[0].strip()}"]=[1,0,0,0,0,0,0]      #   [over0, medan1, runs2, Wickets3, NB4, WD5, ECO6]
    elif "wide" in temp[1]:
        pass
    elif "byes" in temp[1]:
        if "FOUR" in temp[2]:
            india_byes+=4
            pakistan_bowlers[f"{current_ball[0].strip()}"][0]+=1
        elif "1" in temp[2]:
            india_byes+=1
            pakistan_bowlers[f"{current_ball[0].strip()}"][0]+=1
        elif "2" in temp[2]:
            india_byes+=2
            pakistan_bowlers[f"{current_ball[0].strip()}"][0]+=1
        elif "3" in temp[2]:
            india_byes+=3
            pakistan_bowlers[f"{current_ball[0].strip()}"][0]+=1
        elif "4" in temp[2]:
            india_byes+=4
            pakistan_bowlers[f"{current_ball[0].strip()}"][0]+=1
        elif "5" in temp[2]:
            india_byes+=5
            pakistan_bowlers[f"{current_ball[0].strip()}"][0]+=1
    else:
        pakistan_bowlers[f"{current_ball[0].strip()}"][0]+=1
    
    if f"{current_ball[1].strip()}" not in india_bats.keys() and temp[1]!="wide":
        india_bats[f"{current_ball[1].strip()}"]=[0,1,0,0,0]    #   [runs, ball, 4s, 6s, sr]
    elif "wide" in temp[1] :
        pass
    else:
        india_bats[f"{current_ball[1].strip()}"][1]+=1
    

    if "out" in temp[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][3]+=1
        
        if "Bowled" in temp[1].split("!!")[0]:
            out_india_bat[f"{current_ball[1].strip()}"]=("b" + current_ball[0])
        elif "Caught" in temp[1].split("!!")[0]:
            w=(temp[1].split("!!")[0]).split("by")
            out_india_bat[f"{current_ball[1].strip()}"]=("c" + w[1] +" b " + current_ball[0])
        elif "Lbw" in temp[1].split("!!")[0]:
            out_india_bat[f"{current_ball[1].strip()}"]=("lbw  b "+current_ball[0])

    
    
    if "no run" in temp[1] or "out" in temp[1] :
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=0
        india_bats[f"{current_ball[1].strip()}"][0]+=0
    elif "1 run" in temp[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=1
        india_bats[f"{current_ball[1].strip()}"][0]+=1
    elif "2 runs" in temp[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=2
        india_bats[f"{current_ball[1].strip()}"][0]+=2
    elif "3 runs" in temp[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=3
        india_bats[f"{current_ball[1].strip()}"][0]+=3
    elif "4 runs" in temp[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=4
        india_bats[f"{current_ball[1].strip()}"][0]+=4
    elif "FOUR" in temp[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=4
        india_bats[f"{current_ball[1].strip()}"][0]+=4
        india_bats[f"{current_ball[1].strip()}"][2]+=1
    elif "SIX" in temp[1]:
        pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=6
        india_bats[f"{current_ball[1].strip()}"][0]+=6
        india_bats[f"{current_ball[1].strip()}"][3]+=1
    elif "wide" in temp[1]:
        if "wides" in temp[1]:
            pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=int(temp[1][1])
            pakistan_bowlers[f"{current_ball[0].strip()}"][5]+=int(temp[1][1])
        else:
            pakistan_bowlers[f"{current_ball[0].strip()}"][2]+=1
            pakistan_bowlers[f"{current_ball[0].strip()}"][5]+=1


for value in india_bats.values():
    value[-1]=round((value[0]/value[1])*100 , 2)

for value in pakistan_bats.values():
    value[-1]=round((value[0]/value[1])*100 , 2)

for value in india_bowlers.values():
    if value[0]%6==0:
        value[0] = value[0]//6
    else:
        value[0] = (value[0]//6) + (value[0]%6)/10

for value in pakistan_bowlers.values():
    if value[0]%6==0:
        value[0] = value[0]//6
    else:
        value[0] = (value[0]//6) + (value[0]%6)/10

for value in india_bowlers.values():
    t=str(value[0])
    if "." in t:
        balls = int(t[0])*6 + int(t[2])
        value[-1]=round((value[2]/balls)*6,1)
    else:
        value[-1] = round((value[2]/value[0]) ,1) 


for value in pakistan_bowlers.values():
    t=str(value[0])
    if "." in t:
        balls = int(t[0])*6 + int(t[2])
        value[-1]=round((value[2]/balls)*6,1)
    else:
        value[-1] = round((value[2]/value[0]) ,1)


    # pakistan batting

names_pakistan_batters=[]
for key in pakistan_bats.keys():
    names_pakistan_batters.append(key)


for i in range(len(pakistan_bats)):
    spreadsheet.cell(5+i,1).value = names_pakistan_batters[i]
    spreadsheet.cell(5+i,5).value = pakistan_bats[names_pakistan_batters[i]][0]
    spreadsheet.cell(5+i,6).value = pakistan_bats[names_pakistan_batters[i]][1]
    spreadsheet.cell(5+i,7).value = pakistan_bats[names_pakistan_batters[i]][2]
    spreadsheet.cell(5+i,8).value = pakistan_bats[names_pakistan_batters[i]][3]
    spreadsheet.cell(5+i,9).value = pakistan_bats[names_pakistan_batters[i]][4]
    if names_pakistan_batters[i] not in out_pakistan_bat:
        spreadsheet.cell(5+i,3).value = "not out"
    else:
        spreadsheet.cell(5+i,3).value=out_pakistan_bat[names_pakistan_batters[i]]

spreadsheet.cell(3,1).value = "BATTERS"
spreadsheet["E3"] = "RUNS"
spreadsheet["F3"] = "BALLS"
spreadsheet["G3"] = " 4s "
spreadsheet["H3"] = " 6s "
spreadsheet["I3"] = "  SR  "

    # india bowling

spreadsheet["A18"] = "BOWLER"
spreadsheet["C18"] = "OVER"
spreadsheet["D18"] = "MAIDEN"
spreadsheet["E18"] = "RUNS"
spreadsheet["F18"] = "WICKET"
spreadsheet["G18"] = "NO-BALL"
spreadsheet["H18"] = "WIDE"
spreadsheet["I18"] = "ECONOMY"

names_pakistan_bowlers=[]
for key in pakistan_bowlers.keys():
    names_pakistan_bowlers.append(key)

for i in range(len(pakistan_bowlers)):
    spreadsheet.cell(42+i,1).value = names_pakistan_bowlers[i]
    spreadsheet.cell(42+i,3).value = pakistan_bowlers[names_pakistan_bowlers[i]][0]
    spreadsheet.cell(42+i,4).value = pakistan_bowlers[names_pakistan_bowlers[i]][1]
    spreadsheet.cell(42+i,5).value = pakistan_bowlers[names_pakistan_bowlers[i]][2]
    spreadsheet.cell(42+i,6).value = pakistan_bowlers[names_pakistan_bowlers[i]][3]
    spreadsheet.cell(42+i,7).value = pakistan_bowlers[names_pakistan_bowlers[i]][4]
    spreadsheet.cell(42+i,8).value = pakistan_bowlers[names_pakistan_bowlers[i]][5]
    spreadsheet.cell(42+i,9).value = pakistan_bowlers[names_pakistan_bowlers[i]][6]
    pakistan_bowlers_total+=pakistan_bowlers[names_pakistan_bowlers[i]][2]
    india_fall_of_wickets+=pakistan_bowlers[names_pakistan_bowlers[i]][3]

    # india batting
spreadsheet.cell(11+len(pakistan_bats)+len(pakistan_bowlers),1).value = "# INDIA"
spreadsheet.cell(11+len(pakistan_bats)+len(pakistan_bowlers),2).value = " INNINGS"

names_india_batters=[]
for key in india_bats.keys():
    names_india_batters.append(key)


for i in range(len(india_bats)):
    spreadsheet.cell(31+i,1).value = names_india_batters[i]
    spreadsheet.cell(31+i,5).value = india_bats[names_india_batters[i]][0]
    spreadsheet.cell(31+i,6).value = india_bats[names_india_batters[i]][1]
    spreadsheet.cell(31+i,7).value = india_bats[names_india_batters[i]][2]
    spreadsheet.cell(31+i,8).value = india_bats[names_india_batters[i]][3]
    spreadsheet.cell(31+i,9).value = india_bats[names_india_batters[i]][4]

    if names_india_batters[i] not in out_india_bat:
        spreadsheet.cell(31+i,3).value = "not out"
    else:
        spreadsheet.cell(31+i,3).value=out_india_bat[names_india_batters[i]]

spreadsheet["A29"] = "BATTERS"
spreadsheet["E29"] = "RUNS"
spreadsheet["F29"] = "BALLS"
spreadsheet["G29"] = " 4s "
spreadsheet["H29"] = " 6s "
spreadsheet["I29"] = "  SR  "

    # india bowling

spreadsheet["A40"] = "BOWLER"
spreadsheet["C40"] = "OVER"
spreadsheet["D40"] = "MAIDEN"
spreadsheet["E40"] = "RUNS"
spreadsheet["F40"] = "WICKET"
spreadsheet["G40"] = "NO-BALL"
spreadsheet["H40"] = "WIDE"
spreadsheet["I40"] = "ECONOMY"

names_india_bowlers=[]
for key in india_bowlers.keys():
    names_india_bowlers.append(key)

for i in range(len(india_bowlers)):

    spreadsheet.cell(20+i,1).value = names_india_bowlers[i]
    spreadsheet.cell(20+i,3).value = india_bowlers[names_india_bowlers[i]][0]
    spreadsheet.cell(20+i,4).value = india_bowlers[names_india_bowlers[i]][1]
    spreadsheet.cell(20+i,5).value = india_bowlers[names_india_bowlers[i]][2]
    spreadsheet.cell(20+i,6).value = india_bowlers[names_india_bowlers[i]][3]
    spreadsheet.cell(20+i,7).value = india_bowlers[names_india_bowlers[i]][4]
    spreadsheet.cell(20+i,8).value = india_bowlers[names_india_bowlers[i]][5]
    spreadsheet.cell(20+i,9).value = india_bowlers[names_india_bowlers[i]][6]
    india_bowlers_total+=india_bowlers[names_india_bowlers[i]][2]
    pakistan_fall_of_wickets+=india_bowlers[names_india_bowlers[i]][3]

india_total_score=india_bowlers_total+pakistan_byes+1
pakistan_total_score = pakistan_bowlers_total+india_byes-1


spreadsheet["E27"] = " "+str(india_total_score) +" - " + str(india_fall_of_wickets)
spreadsheet["F27"] = str(overs_india)
Eone=" "+str(pakistan_total_score) +" - " + str(pakistan_fall_of_wickets)
Fone = str(overs_pakistan)

wb.save("Scoreboard.xlsx")

df = pd.read_excel('Scoreboard.xlsx')

df = df.set_axis(['PAKISTAN', ' INNINGS'] + [" "," ",Eone,Fone," "," "," "], axis='columns')

df.to_csv('Scorecard.csv',index=False)

import os
try:
    os.path.exists("Scoreboard.xlsx") 
    os.remove("Scoreboard.xlsx")        # deleting output excel
except:
    print("Extra created file does not exist")


#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
